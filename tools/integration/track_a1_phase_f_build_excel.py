"""트랙 A1 Phase F — 통합 회귀 매트릭스 엑셀 산출.

목적:
- Phase C/E 검증 + 기존 누적 매트릭스 자산을 한 엑셀 보고서로 통합
- 5계정 × 18 화면 × 모든 액션 의 정신을 살려, 모든 매트릭스 PASS/FAIL 일목요연

수집 매트릭스 (user_mig/*.json):
- ADMIN_CONSOLE_SMOKE.json — 18 페이지 admin smoke
- USER_SCENARIO_E2E.json — 9 단계 사용자 시나리오
- I18N_LOCALE_MATRIX.json — 3 locale × 5 페이지
- TRACK_A1_PHASE_C.json — 7 단계 Users mutation
- PHASE5_VALIDATION.json — 4 routing scenario (있으면)
"""
import io
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1].parent
USER_MIG = ROOT / "user_mig"
OUT = USER_MIG / f"A1_PHASE_F_REGRESSION_MATRIX_{datetime.now():%Y%m%d}.xlsx"


# ── 스타일 ──────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_data_row(ws, row: int, last_col: int, status: str) -> None:
    fill = PASS_FILL if status == "PASS" else (FAIL_FILL if status == "FAIL" else None)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load_json(name: str) -> dict | None:
    p = USER_MIG / name
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def main() -> None:
    matrices = {
        "관리자 콘솔 smoke (18 페이지)": load_json("ADMIN_CONSOLE_SMOKE.json"),
        "사용자 시나리오 e2e (9 단계)": load_json("USER_SCENARIO_E2E.json"),
        "i18n 3 locale × 5 페이지": load_json("I18N_LOCALE_MATRIX.json"),
        "Phase C: Users mutation 직접 + 410 (7)": load_json("TRACK_A1_PHASE_C.json"),
        "Phase 5.7 routing matrix (4)": load_json("PHASE5_VALIDATION.json"),
        "Phase E: DocUtil redirect (15)": load_json("DOCUTIL_REDIRECT.json"),
    }

    wb = Workbook()
    wb.remove(wb.active)

    # ─── Cover 시트 ──────────────────────────────────────────────
    cover = wb.create_sheet("Cover")
    cover["A1"] = "IDINO Agent Hub — 통합 회귀 매트릭스"
    cover["A1"].font = Font(bold=True, size=18)
    cover["A2"] = f"트랙 A1 Phase F (2026-06-01) — 통합 작업 완전화 검증"
    cover["A2"].font = Font(size=11, italic=True)
    cover["A4"] = "📦 수집 매트릭스"
    cover["A4"].font = Font(bold=True, size=13)

    row = 5
    cover.cell(row=row, column=1, value="No.")
    cover.cell(row=row, column=2, value="매트릭스")
    cover.cell(row=row, column=3, value="셀")
    cover.cell(row=row, column=4, value="PASS")
    cover.cell(row=row, column=5, value="FAIL")
    cover.cell(row=row, column=6, value="비고")
    style_header_row(cover, row, 6)

    total_pass = total_fail = total_cells = 0
    for idx, (label, data) in enumerate(matrices.items(), start=1):
        row += 1
        if data is None:
            cover.cell(row=row, column=1, value=idx)
            cover.cell(row=row, column=2, value=label)
            cover.cell(row=row, column=3, value="-")
            cover.cell(row=row, column=4, value="-")
            cover.cell(row=row, column=5, value="-")
            cover.cell(row=row, column=6, value="(매트릭스 파일 없음 - 미수집 또는 별도 보관)")
            for c in range(1, 7):
                cover.cell(row=row, column=c).border = BORDER
            continue

        summary = data.get("summary") or {}
        # 형식 1: {summary: {pass, fail, total}}
        # 형식 2: top-level {passed, failed, total} (ADMIN_CONSOLE_SMOKE / PHASE5_VALIDATION)
        p = (summary.get("pass") or summary.get("PASS")
             or data.get("passed") or data.get("PASSED") or 0)
        f = (summary.get("fail") or summary.get("FAIL")
             or data.get("failed") or data.get("FAILED") or 0)
        total = (summary.get("total") or data.get("total") or (p + f))
        total_pass += p
        total_fail += f
        total_cells += total

        cover.cell(row=row, column=1, value=idx)
        cover.cell(row=row, column=2, value=label)
        cover.cell(row=row, column=3, value=total)
        cover.cell(row=row, column=4, value=p)
        cover.cell(row=row, column=5, value=f)
        status = "PASS" if f == 0 else "FAIL"
        cover.cell(row=row, column=6, value="✅ 모두 PASS" if f == 0 else f"❌ FAIL {f}건")
        style_data_row(cover, row, 6, status)

    row += 2
    cover.cell(row=row, column=1, value="합계").font = Font(bold=True)
    cover.cell(row=row, column=3, value=total_cells).font = Font(bold=True)
    cover.cell(row=row, column=4, value=total_pass).font = Font(bold=True)
    cover.cell(row=row, column=5, value=total_fail).font = Font(bold=True)
    cover.cell(
        row=row,
        column=6,
        value=f"운영 결함 {'0' if total_fail == 0 else total_fail} 건",
    ).font = Font(bold=True, color="00B050" if total_fail == 0 else "C00000")
    for c in range(1, 7):
        cover.cell(row=row, column=c).border = BORDER

    autofit(cover, [6, 40, 8, 8, 8, 30])

    # ─── 매트릭스별 상세 시트 ──────────────────────────────────────
    for label, data in matrices.items():
        if data is None:
            continue
        sheet_name = label.split(" ")[0][:31]  # 시트명 31자 제한
        # 중복 방지
        suffix = 1
        base = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = label
        ws["A1"].font = TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        # 결과 추출
        rows = data.get("results") or data.get("matrix") or []

        # 헤더 — 결과 첫 row 의 key 동적 추출
        header_row = 3
        if rows:
            keys = list(rows[0].keys())
            # 결과 표시 우선순위: step/label/locale/path → status → detail
            priority = ["locale", "path", "label", "step", "scenario", "feature", "status", "detail"]
            keys.sort(key=lambda k: (priority.index(k) if k in priority else 999, k))

            for c, key in enumerate(keys, start=1):
                ws.cell(row=header_row, column=c, value=key)
            style_header_row(ws, header_row, len(keys))

            for ridx, r in enumerate(rows, start=header_row + 1):
                for c, key in enumerate(keys, start=1):
                    v = r.get(key)
                    if isinstance(v, (list, dict)):
                        v = json.dumps(v, ensure_ascii=False)[:300]
                    ws.cell(row=ridx, column=c, value=v)
                status = (r.get("status") or "").upper()
                style_data_row(ws, ridx, len(keys), status)

            # 폭 — 마지막 컬럼은 와이드
            widths = [10] * len(keys)
            if "detail" in keys:
                widths[keys.index("detail")] = 60
            if "step" in keys:
                widths[keys.index("step")] = 35
            if "path" in keys:
                widths[keys.index("path")] = 30
            if "label" in keys:
                widths[keys.index("label")] = 25
            autofit(ws, widths)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\n[saved] {OUT}")
    print(f"        총 {total_cells} 셀  PASS={total_pass} FAIL={total_fail}")
    print(f"        시트: {wb.sheetnames}")


if __name__ == "__main__":
    main()
