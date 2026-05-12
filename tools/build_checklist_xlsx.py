"""트랙 #73 — 테스트 체크리스트 엑셀 생성기.

입력:
- `docs/TEST_CASES.md` (카탈로그 — 사람이 읽는 정본)
- `tools/probe_all_result.json` (라이브 테스트 결과)

출력:
- `docs/TEST_CHECKLIST.xlsx`
  - 11 시트 (A~K 영역)
  - 헤더 색상 강조 (파랑)
  - PASS=녹색 / FAIL=빨강 / SKIP=회색 conditional formatting
  - 위험도 컬럼 색상 (안전=초록 / 중간=주황 / 위험=빨강)
  - 마지막 시트 "Summary" — 영역별 집계
"""
from __future__ import annotations
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).resolve().parent.parent
RESULT_JSON = ROOT / "tools" / "probe_all_result.json"
OUTPUT_XLSX = ROOT / "docs" / "TEST_CHECKLIST.xlsx"

# ==== 카탈로그 정의 ====
# (id, scenario, precond, input_summary, expected, risk, automatable)
# risk: 안전 / 중간 / 위험
# automatable: YES / PARTIAL / NO

CATALOG: dict[str, list[tuple[str, str, str, str, str, str, str]]] = {
    "A_인증": [
        ("A-01", "admin 로그인 (정상 자격증명)", "운영 가동", "POST /api/auth/login {email:admin@example.com, password:Admin123!}", "200 + JWT (~555 chars)", "안전", "YES"),
        ("A-02", "로그인 실패 (잘못된 비밀번호)", "운영 가동", "POST /api/auth/login {password:wrong!}", "400/401/422", "안전", "YES"),
        ("A-03", "anonymous endpoint 차단", "운영 가동", "GET /api/agents (Authorization 없음)", "401", "안전", "YES"),
        ("A-04", "JWT 발급 → /api/agents 200", "A-01 토큰", "GET /api/agents Authorization: Bearer <jwt>", "200 + ≥15 Agent", "안전", "YES"),
        ("A-05", "JWT refresh", "refresh token 인프라", "POST /api/auth/refresh", "200 + 새 JWT", "중간", "PARTIAL"),
        ("A-06", "로그아웃", "A-01 토큰", "POST /api/auth/logout (JSON body)", "200/204/415", "안전", "YES"),
        ("A-07", "추가 admin (hslee@idino.co.kr) 로그인", "자격증명 보유", "POST /api/auth/login", "200 + JWT", "안전", "PARTIAL"),
    ],
    "B_Agent관리": [
        ("B-01", "GET /api/agents (시드 15)", "admin JWT", "GET /api/agents", "200 + ≥15건", "안전", "YES"),
        ("B-02", "GET /api/agents/{id}", "B-01 ID", "GET /api/agents/{agentId}", "200 + 단건", "안전", "YES"),
        ("B-03", "POST /api/agents (신규)", "admin JWT", "POST /api/agents {agentCode, ...}", "201 + 신규 ID", "위험", "YES"),
        ("B-04", "PUT /api/agents/{id}", "신규 Agent", "PUT /api/agents/{id}", "200 + 반영", "위험", "YES"),
        ("B-05", "LlmRouting 전환", "신규 Agent", "PUT llmRouting (External/Internal/Hybrid)", "변경 반영", "위험", "YES"),
        ("B-06", "KnowledgeBaseSource 전환", "신규 Agent", "PUT knowledgeBaseSource (DocUtil/AgentHub)", "KB 소스 변경", "위험", "YES"),
        ("B-07", "EnableRag 토글", "신규 Agent", "PUT enableRag", "RAG flag 변경", "위험", "YES"),
    ],
    "C_ApiKey": [
        ("C-01", "GET /api/api-keys 목록", "admin JWT", "GET /api/api-keys", "200 + 배열", "안전", "YES"),
        ("C-02", "POST /api/api-keys 발급", "admin JWT", "POST /api/api-keys {name, scopes, agentId}", "201 + fullKey 1회", "위험", "YES"),
        ("C-03", "ApiKey 회수 (isActive=false)", "C-02 키", "PUT /api/api-keys/{id} isActive:false", "200 + 이후 호출 401", "위험", "YES"),
        ("C-04", "X-API-Key /v1/* 호출", "유효 ApiKey", "GET /v1/models X-API-Key: ak-...", "200 + 모델", "안전", "YES"),
    ],
    "D_OpenAI호환": [
        ("D-01", "GET /v1/models", "유효 ApiKey", "X-API-Key 헤더", "200 + 카탈로그(15)", "안전", "YES"),
        ("D-02", "POST /v1/chat/completions sync", "유효 ApiKey", "{model:docutil-rag-chat, messages:[...]}", "200 + assistant 응답", "위험", "YES"),
        ("D-03", "POST /v1/chat/completions stream", "유효 ApiKey", "{stream:true}", "200 + SSE + [DONE]", "위험", "YES"),
        ("D-04", "POST /v1/embeddings", "유효 ApiKey", "{model:text-embedding-3-small, input:...}", "200 + 1536D 벡터", "위험", "YES"),
        ("D-05", "POST /v1/images/generations", "유효 ApiKey", "{model:dall-e-3, prompt}", "200 + url/b64_json", "위험", "YES"),
        ("D-06", "Internal 라우팅 (Nexus)", "Internal Agent", "/v1/chat/completions", "200 + Nexus 응답", "위험", "PARTIAL"),
        ("D-07", "Hybrid 라우팅 PII", "Hybrid Agent", "PII 포함 입력", "Internal 라우팅 (로그)", "위험", "PARTIAL"),
    ],
    "E_채팅": [
        ("E-01", "POST /api/chat/send RAG", "RAG Agent + admin JWT", "{agentId, message}", "200 + 한국어 + context", "위험", "YES"),
        ("E-02", "SignalR negotiate", "admin JWT", "POST /hubs/notification/negotiate", "200 + connectionToken", "중간", "PARTIAL"),
        ("E-03", "게스트 Rate Limit", "미인증", "21회 호출", "20+1 (429)", "중간", "YES"),
        ("E-04", "PII 입력 차단", "PII Agent", "주민번호 입력", "거절 또는 마스킹", "중간", "YES"),
    ],
    "F_Tool_Workflow": [
        ("F-01", "GET /api/tools", "admin JWT", "GET /api/tools", "200 또는 404", "안전", "YES"),
        ("F-02", "POST /api/tools (C#/Script/API)", "admin JWT", "POST /api/tools {toolType, ...}", "201 + 등록", "위험", "YES"),
        ("F-03", "Tool 실행 (Tool Calling)", "Tool + Agent", "chat 유도", "tool 호출 + 결과", "위험", "YES"),
        ("F-04", "GET /api/workflows", "admin JWT", "GET /api/workflows", "200 또는 404", "안전", "YES"),
        ("F-05", "Workflow 실행", "Workflow 정의", "POST /api/workflows/{id}/run", "200 + 실행", "위험", "PARTIAL"),
    ],
    "G_AdminBFF_DocUtil": [
        ("G-01a", "users anonymous 차단", "-", "GET /api/admin/docutil/users", "401", "안전", "YES"),
        ("G-01b", "users admin read", "admin JWT", "GET /api/admin/docutil/users", "200/404/502", "중간", "YES"),
        ("G-02a", "departments anonymous", "-", "GET /api/admin/docutil/departments", "401", "안전", "YES"),
        ("G-02b", "departments admin read", "admin JWT", "GET /api/admin/docutil/departments", "200/404/502", "중간", "YES"),
        ("G-03a", "projects anonymous", "-", "GET /api/admin/docutil/projects", "401", "안전", "YES"),
        ("G-03b", "projects admin read", "admin JWT", "GET /api/admin/docutil/projects", "200/404/502", "중간", "YES"),
        ("G-04a", "dashboard anonymous", "-", "GET /api/admin/docutil/dashboard/summary", "401", "안전", "YES"),
        ("G-04b", "dashboard admin read", "admin JWT", "GET /api/admin/docutil/dashboard/summary", "200/404/502", "중간", "YES"),
        ("G-05a", "audit-logs anonymous", "-", "GET /api/admin/docutil/audit-logs", "401", "안전", "YES"),
        ("G-05b", "audit-logs admin read", "admin JWT", "GET /api/admin/docutil/audit-logs", "200/404/502", "중간", "YES"),
        ("G-06a", "search-scopes anonymous", "-", "GET /api/admin/docutil/search-scopes", "401", "안전", "YES"),
        ("G-06b", "search-scopes admin read", "admin JWT", "GET /api/admin/docutil/search-scopes", "200/404/502", "중간", "YES"),
        ("G-07a", "evaluation anonymous", "-", "GET /api/admin/docutil/evaluation", "401", "안전", "YES"),
        ("G-07b", "evaluation admin read", "admin JWT", "GET /api/admin/docutil/evaluation", "200/404/502", "중간", "YES"),
        ("G-08a", "faq anonymous", "-", "GET /api/admin/docutil/faq", "401", "안전", "YES"),
        ("G-08b", "faq admin read", "admin JWT", "GET /api/admin/docutil/faq", "200/404/502", "중간", "YES"),
        ("G-09a", "reports anonymous", "-", "GET /api/admin/docutil/reports", "401", "안전", "YES"),
        ("G-09b", "reports admin read", "admin JWT", "GET /api/admin/docutil/reports", "200/404/410/502", "중간", "YES"),
        ("G-10a", "templates anonymous", "-", "GET /api/admin/docutil/templates", "401", "안전", "YES"),
        ("G-10b", "templates admin read", "admin JWT", "GET /api/admin/docutil/templates", "200/404/502", "중간", "YES"),
        ("G-11a", "api-keys (deprecate) anonymous", "-", "GET /api/admin/docutil/api-keys", "401", "안전", "YES"),
        ("G-11b", "api-keys admin read", "admin JWT", "GET /api/admin/docutil/api-keys", "200/404/502", "중간", "YES"),
        ("G-12a", "agents (DocUtil 챗봇) anonymous", "-", "GET /api/admin/docutil/agents", "401", "안전", "YES"),
        ("G-12b", "agents admin read", "admin JWT", "GET /api/admin/docutil/agents", "200/404/502", "중간", "YES"),
        ("G-13a", "documents-v2 anonymous", "-", "GET /api/admin/docutil/documents-v2", "401", "안전", "YES"),
        ("G-13b", "documents-v2 admin read", "admin JWT", "GET /api/admin/docutil/documents-v2", "200/404/502", "중간", "YES"),
    ],
    "H_Analytics_Quota_Audit": [
        ("H-01", "GET /api/analytics/usage", "admin JWT", "GET /api/analytics/usage", "200 또는 404", "안전", "YES"),
        ("H-02", "GET /api/quota", "admin JWT", "GET /api/quota", "200 또는 404", "안전", "YES"),
        ("H-03", "GET /api/audit", "admin JWT", "GET /api/audit", "200 또는 404", "안전", "YES"),
    ],
    "I_DocUtil_사용자": [
        ("I-01", "DocUtil 로그인", "DocUtil 자격증명", "POST /api/v1/auth/login", "200 + DocUtil JWT", "안전", "YES"),
        ("I-02", "DocUtil 챗봇 (AgentHub 위임)", "I-01 토큰", "POST /api/v1/chat", "200 + LLM 응답", "위험", "YES"),
        ("I-03", "DocUtil /api/v1/search", "I-01 토큰", "POST /api/v1/search {query}", "200 + 청크", "중간", "YES"),
        ("I-04", "DocUtil 문서 업로드", "I-01 토큰", "POST /api/v1/documents (multipart)", "201 + 문서 ID + 인덱싱", "위험", "YES"),
        ("I-05", "DocUtil 보고서 생성", "I-01 토큰", "POST /api/v1/reports/generate", "200 + 보고서", "위험", "PARTIAL"),
    ],
    "J_보안_RateLimit": [
        ("J-01a", "anonymous /api/* 401", "-", "GET /api/agents", "401", "안전", "YES"),
        ("J-01b", "anonymous /api/admin/* 401", "-", "GET /api/admin/docutil/users", "401", "안전", "YES"),
        ("J-02", "User role admin endpoint 403", "비 admin JWT", "GET /api/admin/*", "403", "중간", "PARTIAL"),
        ("J-03", "Rate Limit (per-user 60/min)", "admin JWT × 70회", "70회 호출", "60 + 10(429)", "중간", "YES"),
        ("J-04", "JWT 위조 검증", "-", "Bearer invalid.jwt.token", "401/403", "안전", "YES"),
        ("J-05", "SQL Injection 방어", "-", "?q=' OR 1=1 --", "400 / 안전 처리", "중간", "PARTIAL"),
        ("J-06", "XSS 방어", "-", "<script>alert(1)</script>", "이스케이프", "중간", "PARTIAL"),
        ("J-07", "CORS preflight", "-", "OPTIONS /api/agents", "200/204/405", "안전", "YES"),
    ],
    "K_통합e2e": [
        ("K-01", "Phase 6.5 RAG round-trip", "RAG Agent + DocUtil", "AgentHub /chat/send → DocUtil /api/v1/search", "200 + context", "위험", "PARTIAL"),
        ("K-02", "DocUtil → AgentHub → OpenAI", "DocUtil UI", "메시지 입력", "200 + ApiUsages 기록", "위험", "PARTIAL"),
        ("K-03", "KB 업로드 → AgentBuilder dropdown", "AgentHub admin UI", "BFF 업로드", "DocUtil collection + dropdown", "위험", "NO"),
    ],
}


def main() -> None:
    # 라이브 결과 매핑
    live: dict[str, dict] = {}
    if RESULT_JSON.exists():
        try:
            data = json.loads(RESULT_JSON.read_text(encoding="utf-8"))
            for c in data.get("cases", []):
                live[c["id"]] = c
        except Exception as e:
            print(f"warn: failed to load live results: {e}")

    wb = Workbook()
    wb.remove(wb.active)

    # 스타일
    header_fill = PatternFill("solid", fgColor="305496")  # 진한 파랑
    header_font = Font(bold=True, color="FFFFFF")
    risk_fills = {
        "안전": PatternFill("solid", fgColor="C6EFCE"),  # 연한 초록
        "중간": PatternFill("solid", fgColor="FFEB9C"),  # 연한 주황
        "위험": PatternFill("solid", fgColor="FFC7CE"),  # 연한 빨강
    }
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["ID", "시나리오", "사전조건", "입력", "기대 결과", "위험도", "자동화", "결과", "실측값", "비고", "검증일시"]
    col_widths = [10, 38, 24, 50, 28, 8, 10, 8, 50, 30, 18]

    # 영역별 시트
    summary_rows = []
    for sheet_name, rows in CATALOG.items():
        ws = wb.create_sheet(sheet_name[:31])  # 31자 제한
        # 헤더
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        area_stats = {"PASS": 0, "FAIL": 0, "SKIP": 0, "TOTAL": 0}
        for ri, (rid, scenario, precond, inp, expected, risk, auto) in enumerate(rows, start=2):
            ws.cell(row=ri, column=1, value=rid).alignment = align_center
            ws.cell(row=ri, column=2, value=scenario).alignment = align_left
            ws.cell(row=ri, column=3, value=precond).alignment = align_left
            ws.cell(row=ri, column=4, value=inp).alignment = align_left
            ws.cell(row=ri, column=5, value=expected).alignment = align_left
            risk_cell = ws.cell(row=ri, column=6, value=risk)
            risk_cell.alignment = align_center
            risk_cell.fill = risk_fills.get(risk, PatternFill())
            ws.cell(row=ri, column=7, value=auto).alignment = align_center

            # 라이브 결과 매핑
            r = live.get(rid)
            result_text = r.get("result", "") if r else ""
            actual_status = r.get("actual_status", "") if r else ""
            actual_preview = r.get("actual_preview", "") if r else ""
            duration_ms = r.get("duration_ms", "") if r else ""
            extra = ""
            if rid in ("G-04a", "G-07a") and result_text == "FAIL":
                extra = "[잠재 결함] SPA fallback 200 — 별도 트랙 후보"
            if rid == "A-06" and result_text == "FAIL":
                extra = "415: JSON body 필수, 운영 결함 아님 — TC 보정 권장"
            if rid == "E-02" and result_text == "FAIL":
                extra = "405: GET 차단 정상 — TC 보정 권장 (POST 필요)"
            actual_value = ""
            if result_text:
                actual_value = f"status={actual_status} ({duration_ms}ms) :: {actual_preview[:140]}"

            ws.cell(row=ri, column=8, value=result_text).alignment = align_center
            ws.cell(row=ri, column=9, value=actual_value).alignment = align_left
            ws.cell(row=ri, column=10, value=extra).alignment = align_left
            ws.cell(row=ri, column=11, value=data.get("finished_at", "") if r else "").alignment = align_center

            for ci in range(1, 12):
                ws.cell(row=ri, column=ci).border = border

            if result_text in area_stats:
                area_stats[result_text] += 1
            area_stats["TOTAL"] += 1

        # conditional formatting — 결과 컬럼 (H)
        last_row = len(rows) + 1
        rng = f"H2:H{last_row}"
        pass_fill = PatternFill("solid", fgColor="92D050")
        fail_fill = PatternFill("solid", fgColor="FF6B6B")
        skip_fill = PatternFill("solid", fgColor="D9D9D9")
        ws.conditional_formatting.add(rng, FormulaRule(formula=['$H2="PASS"'], fill=pass_fill))
        ws.conditional_formatting.add(rng, FormulaRule(formula=['$H2="FAIL"'], fill=fail_fill))
        ws.conditional_formatting.add(rng, FormulaRule(formula=['$H2="SKIP"'], fill=skip_fill))

        # freeze 헤더
        ws.freeze_panes = "A2"

        summary_rows.append((sheet_name, area_stats))

    # Summary 시트
    ws_sum = wb.create_sheet("Summary", 0)  # 맨 앞에
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 10
    ws_sum.column_dimensions["D"].width = 10
    ws_sum.column_dimensions["E"].width = 10

    ws_sum.cell(row=1, column=1, value="IDINO Agent Hub — 테스트 체크리스트 요약").font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:E1")
    ws_sum.cell(row=2, column=1, value=f"생성일시: {(json.loads(RESULT_JSON.read_text(encoding='utf-8'))).get('finished_at', '-') if RESULT_JSON.exists() else '-'}")
    ws_sum.merge_cells("A2:E2")

    sum_headers = ["영역", "총건수", "PASS", "FAIL", "SKIP"]
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=4, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border

    total = {"PASS": 0, "FAIL": 0, "SKIP": 0, "TOTAL": 0}
    for ri, (name, stats) in enumerate(summary_rows, start=5):
        ws_sum.cell(row=ri, column=1, value=name).alignment = align_left
        ws_sum.cell(row=ri, column=2, value=stats["TOTAL"]).alignment = align_center
        ws_sum.cell(row=ri, column=3, value=stats["PASS"]).alignment = align_center
        ws_sum.cell(row=ri, column=4, value=stats["FAIL"]).alignment = align_center
        ws_sum.cell(row=ri, column=5, value=stats["SKIP"]).alignment = align_center
        for ci in range(1, 6):
            ws_sum.cell(row=ri, column=ci).border = border
        for k in total:
            total[k] += stats.get(k, 0)

    total_row = 5 + len(summary_rows)
    ws_sum.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
    ws_sum.cell(row=total_row, column=2, value=total["TOTAL"]).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=3, value=total["PASS"]).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=4, value=total["FAIL"]).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=5, value=total["SKIP"]).font = Font(bold=True)
    for ci in range(1, 6):
        ws_sum.cell(row=total_row, column=ci).border = border
        ws_sum.cell(row=total_row, column=ci).fill = PatternFill("solid", fgColor="FFF2CC")

    # FAIL 상세 + SKIP 사유 노트
    note_row = total_row + 2
    ws_sum.cell(row=note_row, column=1, value="FAIL 상세 (트랙 #74)").font = Font(bold=True)
    note_row += 1
    fails = [
        ("A-06", "POST /api/auth/logout → 415", "JSON body 필수, TC 보정 권장. 운영 결함 아님."),
        ("E-02", "GET /hubs/notification/negotiate → 405", "SignalR negotiate는 POST. TC 보정 권장."),
        ("G-04a", "GET /api/admin/docutil/dashboard/summary anonymous → 200 (text/html)", "잠재 결함 — SPA fallback 노출. 별도 트랙 후보."),
        ("G-07a", "GET /api/admin/docutil/evaluation anonymous → 200 (text/html)", "잠재 결함 — SPA fallback 노출. 별도 트랙 후보."),
    ]
    ws_sum.cell(row=note_row, column=1, value="ID").font = Font(bold=True)
    ws_sum.cell(row=note_row, column=2, value="현상").font = Font(bold=True)
    ws_sum.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=3)
    ws_sum.cell(row=note_row, column=4, value="분석").font = Font(bold=True)
    ws_sum.merge_cells(start_row=note_row, start_column=4, end_row=note_row, end_column=5)
    note_row += 1
    for rid, desc, analysis in fails:
        ws_sum.cell(row=note_row, column=1, value=rid)
        ws_sum.cell(row=note_row, column=2, value=desc).alignment = align_left
        ws_sum.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=3)
        ws_sum.cell(row=note_row, column=4, value=analysis).alignment = align_left
        ws_sum.merge_cells(start_row=note_row, start_column=4, end_row=note_row, end_column=5)
        note_row += 1

    note_row += 1
    ws_sum.cell(row=note_row, column=1, value="SKIP 35건 사유 분류").font = Font(bold=True)
    note_row += 1
    skip_reasons = [
        ("LLM 비용", "D-02~D-05, D-06~D-07, E-01/E-03/E-04, I-02/I-05, K-01~K-03 (18건)"),
        ("운영 mutation 위험", "B-03~B-07, C-02/C-03, F-02/F-03/F-05, I-04, K-03 (13건)"),
        ("추가 인증 정보 미보유", "A-05, A-07, I-01/I-03, J-02 (5건)"),
        ("외부 부하 / 환경 의존", "J-03/J-05/J-06 (3건)"),
    ]
    for cat, ids in skip_reasons:
        ws_sum.cell(row=note_row, column=1, value=cat)
        ws_sum.cell(row=note_row, column=2, value=ids).alignment = align_left
        ws_sum.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=5)
        note_row += 1

    ws_sum.freeze_panes = "A5"

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"saved: {OUTPUT_XLSX} ({OUTPUT_XLSX.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
