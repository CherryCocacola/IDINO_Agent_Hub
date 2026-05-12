"""트랙 #84 — track84_results.json 결과를 두 엑셀에 반영.

대상:
  1. docs/TEST_CHECKLIST.xlsx       (79 케이스, A_인증/B_Agent관리/... 시트)
  2. docs/TEST_CHECKLIST_FULL.xlsx  (479 케이스, 화면별 시트 + SP_* 인용)

매핑:
  - 79 케이스: ID 단위 직접 매칭 (D-02, B-03, F-02, J-04, ...)
  - 479 케이스: SP-* 인용은 그대로 두고, 추가 케이스가 없으므로 79 케이스 갱신 분만 반영
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "tools" / "track84_results.json"
XLSX_79 = ROOT / "docs" / "TEST_CHECKLIST.xlsx"
XLSX_FULL = ROOT / "docs" / "TEST_CHECKLIST_FULL.xlsx"


def main() -> None:
    if not RESULTS.exists():
        print(f"[merge] results not found: {RESULTS}")
        return

    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    # id → result mapping
    id_to_result: dict[str, dict] = {}
    for r in data["results"]:
        id_to_result[r["id"]] = r

    finished_at = data.get("finished_at", "")

    # ── 79 케이스 엑셀 ─────────────────────────────────────────────────────
    if XLSX_79.exists():
        wb = load_workbook(str(XLSX_79))
        updates_79 = 0
        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary":
                continue
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                cid = ws.cell(row=row, column=1).value  # ID
                if not isinstance(cid, str):
                    continue
                r = id_to_result.get(cid)
                if not r:
                    continue
                # 컬럼: 1=ID, 2=시나리오, 3=사전조건, 4=입력, 5=기대결과, 6=위험도, 7=자동화, 8=결과, 9=실측값, 10=비고, 11=검증일시
                ws.cell(row=row, column=8, value=r["result"])
                ws.cell(row=row, column=9,
                        value=f"{r.get('actual_preview', '')[:140]} ({r.get('duration_ms', 0)}ms)")
                # 기존 비고 보존 + 트랙 #84 갱신 표시
                cur_note = ws.cell(row=row, column=10).value or ""
                new_note = f"[트랙 #84 갱신] {r.get('note', '')[:120]}"
                ws.cell(row=row, column=10, value=new_note if not cur_note else f"{new_note} | {cur_note[:80]}")
                ws.cell(row=row, column=11, value=finished_at)
                updates_79 += 1

        # Summary 시트 재집계 (필요 시 COUNTIF 수식이 자동 갱신)
        wb.save(str(XLSX_79))
        print(f"[merge] {XLSX_79.name}: {updates_79} cells updated")

    # ── 479 케이스 엑셀 (TEST_CHECKLIST_FULL.xlsx) ─────────────────────────
    # full 엑셀은 SP_* 시나리오 인용 + 화면별 시트 구조 — 79 케이스 ID 가 직접 매칭되는
    # 셀이 없으므로, Cover 시트에 트랙 #84 결과 요약 추가.
    if XLSX_FULL.exists():
        wb = load_workbook(str(XLSX_FULL))
        # Cover 시트에 트랙 #84 요약 추가
        if "00_Cover" in wb.sheetnames:
            ws = wb["00_Cover"]
            # 마지막 행 찾기
            last_row = ws.max_row
            # 트랙 #84 메타 추가
            from openpyxl.styles import Font
            bold = Font(name="맑은 고딕", size=10, bold=True)
            normal = Font(name="맑은 고딕", size=10)

            track84_rows = [
                ("", ""),
                ("=== 트랙 #84 갱신 (2026-05-12) ===", ""),
                ("트랙 #84", "SKIP 전수 진행 — 자격증명 의존 22건만 SKIP 유지"),
                ("진행 결과",
                 f"PASS={data['summary']['PASS']} / FAIL={data['summary']['FAIL']} / SKIP={data['summary']['SKIP']} / TOTAL={data['summary']['TOTAL']}"),
                ("LLM 비용", f"~${data.get('cost_estimate_usd', 0)} (D-02/03/07/E-01/04 ~5회 호출)"),
                ("운영 영향",
                 "0건 — mutation cycle 7건 모두 cleanup verified (B-03 Agent / F-02 Tool / B-04~07 PUT 원복 / 임시 ApiKey 회수)"),
                ("운영 결함 발견",
                 "Tools/Workflows axios 직접 사용 → JWT 미부착 (anti-patterns.md §11 위반) — 수정 완료"),
            ]
            for i, (k, v) in enumerate(track84_rows, start=1):
                ws.cell(row=last_row + i, column=1, value=k).font = bold if k else normal
                ws.cell(row=last_row + i, column=2, value=v).font = normal

        wb.save(str(XLSX_FULL))
        print(f"[merge] {XLSX_FULL.name}: Cover 시트에 트랙 #84 요약 추가")


if __name__ == "__main__":
    main()
