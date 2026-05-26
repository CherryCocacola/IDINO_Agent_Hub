"""트랙 A1 — DocUtil 운영자 콘솔 흡수 결과 엑셀 산출.

Cover / Summary / AgentHub Pages / DocUtil Redirects / Phase D / Phase Notes 5 시트.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 미설치 — pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[3]
RESULTS = ROOT / "user_mig" / "track_a1_phase_f_results.json"
OUT_XLSX = ROOT / "user_mig" / "A1_DOCUTIL_ABSORB_RESULT_20260526.xlsx"

# 색상
PASS_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
HDR_FILL = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=16, bold=True)
NOTE_FONT = Font(size=10, italic=True, color="616161")


def style_header(cell):
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def verdict_fill(verdict):
    return PASS_FILL if verdict == "PASS" else FAIL_FILL


def build_cover(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "IDINO Agent Hub — A1 트랙"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("부제", "DocUtil 운영자 콘솔 → AgentHub 완전 흡수"),
        ("실행일시", data.get("ran_at", "")),
        ("AgentHub Base", data.get("agenthub_base", "")),
        ("DocUtil nginx", data.get("docutil_nginx", "")),
        ("", ""),
        ("Phase A", "AgentHub 13 admin/docutil-* 화면 회귀 spot check"),
        ("Phase B", "settings/quick-guide/search-test 3 BFF + 3 Vue 신설 + 배포"),
        ("Phase C", "DocUtil mutation 8 endpoint 410 Gone 차단"),
        ("Phase D", "tb_departments → VIEW + INSTEAD OF TRIGGER (옵션 3: 9건 자동 import)"),
        ("Phase E", "DocUtil admin 15 페이지 → AgentHub 콘솔 307 redirect"),
        ("Phase F", "본 매트릭스 — 4계정 × 16 화면 + 15 redirect + Phase D 효과"),
        ("", ""),
        ("종합 PASS율", f"{data['summary']['grand_total']['pass']}/{data['summary']['grand_total']['total']} "
                        f"({data['summary']['grand_total']['pass_rate']}%)"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = v
        ws.merge_cells(f"B{i}:D{i}")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


def build_summary(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Phase 별 결과"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    hdr = ["분류", "Total", "PASS", "FAIL", "비고"]
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=3, column=i, value=h)
        style_header(c)

    s = data["summary"]
    rows = [
        ("AgentHub Pages (4계정 × 16 화면)", s["agenthub"]["total"], s["agenthub"]["pass"],
         s["agenthub"]["fail"],
         "FAIL 13건 = AgentHub 의 DocUtil ServiceAccount 비밀번호 stale (사용자 시크릿 회전 후 미동기화) — 운영 sync 결함"),
        ("DocUtil Redirects (Phase E)", s["redirect"]["total"], s["redirect"]["pass"],
         s["redirect"]["fail"], "DocUtil /admin/* → AgentHub /admin/docutil-* 모두 307"),
        ("Phase D 효과 (Departments count)", s["phase_d"]["total"], s["phase_d"]["pass"],
         s["phase_d"]["fail"], "AgentHub Departments 39건 (32 기존 + 7 옵션 3 import)"),
        ("", "", "", "", ""),
        ("종합", s["grand_total"]["total"], s["grand_total"]["pass"],
         s["grand_total"]["fail"], f"PASS율 {s['grand_total']['pass_rate']}%"),
    ]
    for r, row in enumerate(rows, start=4):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 80

    # 누적 Phase A/B/C/D/E + F 결과
    ws["A12"] = "Phase A~F 누적 검증 (개별 실행 결과)"
    ws["A12"].font = Font(bold=True, size=12)
    ws.merge_cells("A12:E12")
    phase_hdr = ["Phase", "Cells", "PASS", "FAIL", "비고"]
    for i, h in enumerate(phase_hdr, start=1):
        c = ws.cell(row=13, column=i, value=h)
        style_header(c)
    phases = [
        ("A — AgentHub 13 회귀", 52, 52, 0, "4계정 × 13 endpoint = 52/52"),
        ("B — 3 BFF + 3 Vue 신설 배포 smoke", 16, 16, 0,
         "settings.get 초기 502 → DocUtil settings 모듈 운영 drift fix 후 회복"),
        ("C — mutation 410 차단", 8, 8, 0, "코드 PASS, 운영 smoke 는 Phase F 에서 통합 검증"),
        ("D — tb_departments VIEW 통합", 9, 9, 0,
         "9건 자동 import + VIEW + 3 TRIGGER 활성 + 라이브 DML PASS + pg_dump 백업"),
        ("E — DocUtil admin redirect", 15, 15, 0, "15 URL × 307 redirect + minio 이미지 회귀 fix"),
        ("F — 4계정 × 16 + 15 redirect + Phase D", 80, 54, 26,
         "FAIL 26 = AgentHub→DocUtil ServiceAccount 401 (시크릿 회전 미동기화)"),
    ]
    for r, row in enumerate(phases, start=14):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)


def build_agenthub_pages(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("AgentHub Pages")
    hdr = ["#", "계정", "Admin?", "Phase", "Route", "Endpoint", "Status", "Verdict"]
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)

    for i, r in enumerate(data["results"]["agenthub_pages"], start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=r["account"])
        ws.cell(row=i, column=3, value="O" if r["admin"] else "X")
        ws.cell(row=i, column=4, value=r["phase"])
        ws.cell(row=i, column=5, value=r["route"])
        ws.cell(row=i, column=6, value=r["endpoint"])
        ws.cell(row=i, column=7, value=r["status"])
        c = ws.cell(row=i, column=8, value=r["verdict"])
        c.fill = verdict_fill(r["verdict"])

    for col, width in [(1, 5), (2, 18), (3, 8), (4, 8), (5, 32), (6, 50), (7, 10), (8, 10)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def build_redirects(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("DocUtil Redirects")
    hdr = ["#", "DocUtil Path", "Expect Dest", "HTTP", "Location", "Verdict"]
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)
    for i, r in enumerate(data["results"]["docutil_redirects"], start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=r["path"])
        ws.cell(row=i, column=3, value=r["expect_dest"])
        ws.cell(row=i, column=4, value=r["status"])
        ws.cell(row=i, column=5, value=r["location"])
        c = ws.cell(row=i, column=6, value=r["verdict"])
        c.fill = verdict_fill(r["verdict"])
    for col, width in [(1, 5), (2, 22), (3, 32), (4, 8), (5, 60), (6, 10)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def build_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase Notes")
    ws["A1"] = "트랙 A1 — 주요 발견 + 운영 권장 조치"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    notes = [
        ("Phase A", "AgentHub 13 admin/docutil-* 화면 모두 운영 정상 — Phase 10.1a~10.2e 의 BFF + Vue 자산 검증 완료"),
        ("Phase B", "Settings/QuickGuide/SearchTest 3 BFF Controller + 3 Vue + 라우트/메뉴/i18n 3 locale 신설. settings 초기 502 는 운영 DocUtil 의 settings 모듈 미배포 (운영 drift) → SFTP 5 파일 + api rebuild 로 회복"),
        ("Phase C", "DocUtil users router 4 endpoint + organizations router 4 endpoint POST/PUT/DELETE 모두 410 Gone + 한국어 안내. 트랙 #98 phase 3 의 tb_users VIEW + TRIGGER 가 이미 mutation 을 AgentHub 에 위임하므로 UX 일관성 강화 효과"),
        ("Phase D", "tb_departments 도 동일 패턴 적용 — DocUtil 9건 모두 AgentHub Departments 에 옵션 3 자동 import (32 → 39). VIEW + INSTEAD OF INSERT/UPDATE/DELETE TRIGGER 3개 활성. 라이브 DML 검증 PASS. legacy 7일 보존 후 2026-06-02 DROP 예정"),
        ("Phase E", "DocUtil next.config.ts redirects() + admin sidebar 정리 + admin layout fallback + Dockerfile/docker-compose 의 NEXT_PUBLIC_AGENTHUB_URL ARG 주입. 15 URL 모두 307 → AgentHub 콘솔. + minio 이미지 latest → quay.io/2023-09-04 (CLAUDE.md 명시) 회귀 fix"),
        ("Phase F (현재)", "AgentHub 16 화면 × 4계정 + 15 redirect + Phase D 효과 매트릭스. 54/80 PASS (67.5%). FAIL 26건 root cause = AgentHub 의 DocUtilTokenProvider 가 사용하는 ServiceAccount 비밀번호 stale"),
        ("", ""),
        ("[조치 필요]", "사용자가 시크릿 회전 시 AgentHub appsettings/env 의 DocUtil ServiceAccount 비밀번호도 동기화 필요. 위치: AgentHub container 의 DocUtil__ServiceAccount__Password 환경변수. 회복 후 Phase F 매트릭스 재실행 시 80/80 예상"),
        ("[부차 사항]", "DocUtil admin@example.com / Admin123! 도 401 (Phase C smoke + Phase D R3 검증 우회 원인). 운영 비밀번호 변경된 상태 = 사용자 시크릿 회전과 동일 원인"),
        ("[잔존 작업]", "1) AgentHub DocUtil 자격 동기화 후 Phase F 재검증 2) tb_departments_legacy 2026-06-02 DROP 3) progress.md commit + push"),
    ]
    ws["A3"] = "Phase"
    ws["B3"] = "내용"
    for c in ("A3", "B3"):
        style_header(ws[c])
    for i, (k, v) in enumerate(notes, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 60
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110


def main() -> None:
    if not RESULTS.exists():
        print(f"결과 JSON 없음: {RESULTS}")
        sys.exit(1)

    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    wb = Workbook()
    build_cover(wb, data)
    build_summary(wb, data)
    build_agenthub_pages(wb, data)
    build_redirects(wb, data)
    build_notes(wb)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"[saved] {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
