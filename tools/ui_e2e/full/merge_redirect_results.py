"""트랙 #83 — anonymous redirect 결과를 -02 케이스에 추가 머지."""
from __future__ import annotations
import io
import json
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from routes_catalog import PROTECTED_ROUTES

from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parents[3] / "docs" / "TEST_CHECKLIST_FULL.xlsx"
RESULTS = Path(__file__).parent / "live_results_redirect.json"

COL_RESULT = 9
COL_ACTUAL = 10
COL_NOTE = 11
COL_VERIFIED_AT = 12


def main() -> None:
    if not RESULTS.exists():
        print(f"results not found: {RESULTS}")
        return
    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    wb = load_workbook(str(XLSX))
    updates = 0
    for r in data["results"]:
        screen = r["screen"]
        path = r["path"]
        safe = screen[:31].replace("/", "_").replace("\\", "_").replace("?", "")
        if safe not in wb.sheetnames:
            print(f"sheet not found: {safe}")
            continue
        ws = wb[safe]
        for row in range(2, ws.max_row + 1):
            cid = ws.cell(row=row, column=1).value
            if isinstance(cid, str) and cid.endswith("-02"):
                # 빈 셀 또는 안 채워진 권한 분기 케이스만 갱신
                cur = ws.cell(row=row, column=COL_RESULT).value
                if not cur:
                    ws.cell(row=row, column=COL_RESULT, value=r["status"])
                    ws.cell(row=row, column=COL_ACTUAL,
                            value=f"final={r['final_url'][:120]}")
                    ws.cell(row=row, column=COL_NOTE, value=r["note"][:300])
                    ws.cell(row=row, column=COL_VERIFIED_AT, value=data["started_at"])
                    updates += 1
                break
    wb.save(str(XLSX))
    print(f"[merge_redirect] updates: {updates}")
    print(f"[merge_redirect] saved: {XLSX}")


if __name__ == "__main__":
    main()
