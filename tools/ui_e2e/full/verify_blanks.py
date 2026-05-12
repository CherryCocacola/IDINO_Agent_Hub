"""트랙 #83 — blank 셀 분석."""
from __future__ import annotations
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parents[3] / "docs" / "TEST_CHECKLIST_FULL.xlsx"


def main() -> None:
    wb = load_workbook(str(XLSX))
    blanks: list[tuple[str, str, str]] = []
    fails: list[tuple[str, str, str]] = []
    for name in wb.sheetnames:
        if name in ("00_Cover", "01_Summary"):
            continue
        ws = wb[name]
        for row in range(2, ws.max_row + 1):
            cid = ws.cell(row=row, column=1).value
            if not cid:
                continue
            res = ws.cell(row=row, column=9).value
            interaction = ws.cell(row=row, column=3).value or ""
            if not res:
                blanks.append((name, str(cid), str(interaction)[:60]))
            elif res == "FAIL":
                fails.append((name, str(cid), str(interaction)[:60]))
    print(f"Blank: {len(blanks)}")
    for s, c, i in blanks[:80]:
        print(f"  [{s}] {c} | {i}")
    print(f"\nFail: {len(fails)}")
    for s, c, i in fails:
        print(f"  [{s}] {c} | {i}")


if __name__ == "__main__":
    main()
