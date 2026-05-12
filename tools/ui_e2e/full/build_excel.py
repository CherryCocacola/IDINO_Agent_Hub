"""트랙 #83 — 전수 화면 체크리스트 엑셀 신설.

산출물: docs/TEST_CHECKLIST_FULL.xlsx (시트 N개 + Summary + Cover)

컬럼:
  ID / 화면 / 인터랙션 / 사전조건 / 입력/액션 / 기대 결과 / 위험도 / 자동화 / 결과 / 실측값 / 비고 / 검증일시 / 스크린샷 경로
"""
from __future__ import annotations
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from routes_catalog import build_all_cases

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parents[3] / "docs" / "TEST_CHECKLIST_FULL.xlsx"

# ── 컬럼 정의 ────────────────────────────────────────────────────────────
COLUMNS = [
    ("ID", 14),
    ("화면", 22),
    ("인터랙션", 50),
    ("사전조건", 30),
    ("입력/액션", 40),
    ("기대 결과", 50),
    ("위험도", 10),
    ("자동화", 10),
    ("결과", 10),
    ("실측값", 38),
    ("비고", 24),
    ("검증일시", 18),
    ("스크린샷 경로", 40),
]

# ── 스타일 ───────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="맑은 고딕", size=9)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header(ws, row: int = 1) -> None:
    for col, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 28


def _write_case(ws, row: int, c) -> None:
    values = [
        c.case_id,
        c.screen,
        c.interaction,
        c.precond,
        c.action,
        c.expected,
        c.risk,
        c.auto,
        "",  # 결과 — 라이브 단계에서 채움
        "",  # 실측값
        c.note,
        "",  # 검증일시
        "",  # 스크린샷 경로
    ]
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = CELL_FONT
        cell.alignment = WRAP
        cell.border = BORDER


def _apply_conditional_formatting(ws, row_count: int) -> None:
    """결과 컬럼(I)에 PASS/FAIL/SKIP 색상."""
    if row_count < 2:
        return
    rng = f"I2:I{row_count}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"PASS"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"FAIL"'],
                   fill=PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"SKIP"'],
                   fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")),
    )


def _build_cover(wb: Workbook, total_cases: int, sheets: list[str]) -> None:
    """표지 시트 — 트랙 #83 메타."""
    ws = wb.create_sheet("00_Cover", 0)
    ws["A1"] = "IDINO Agent Hub — 전수 화면 체크리스트 (트랙 #83)"
    ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True)
    ws.merge_cells("A1:F1")

    rows = [
        ("작업 트랙", "#83 — 전수 화면 체크리스트 + 라이브 e2e"),
        ("작성일", "2026-05-12"),
        ("대상 시스템", "AgentHub Vue 3 (49 routes) + DocUtil Next.js 16 (25 routes)"),
        ("운영 URL", "AgentHub: http://192.168.10.39:64005  |  DocUtil: http://192.168.10.39:8041"),
        ("총 케이스 수", str(total_cases)),
        ("총 시트 수", f"{len(sheets)} (Cover/Summary 포함)"),
        ("운영 영향 정책", "read-only 진입 + 안전 mutation 1 cycle (시나리오 1 인용) 만"),
        ("LLM 비용 정책", "Playground/이미지/PPTX 등은 폼 표시만 — 전송 안 함. 채팅 1회는 시나리오 2 인용"),
        ("DocUtil 자격증명", "미확보 — DocUtil 사용자/관리자 페이지는 SKIP. 자격증명 제공 시 진행 분기"),
        ("관련 commit", "트랙 #75: b7de919 / 3542e33 (시나리오 1~4)"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=i, column=2, value=v).font = Font(name="맑은 고딕", size=10)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    ws.column_dimensions["A"].width = 20
    for col in "BCDEF":
        ws.column_dimensions[col].width = 20
    ws.row_dimensions[1].height = 26


def _build_summary(wb: Workbook, cases) -> None:
    """집계 시트 — 화면별/위험도별/자동화별/결과별 + Summary 수식."""
    ws = wb.create_sheet("01_Summary", 1)
    ws["A1"] = "집계"
    ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True)

    # 시트별 케이스 수 + 결과 집계 수식
    by_sheet: dict[str, int] = {}
    for c in cases:
        by_sheet[c.sheet] = by_sheet.get(c.sheet, 0) + 1

    headers = ["시트", "케이스 수", "PASS", "FAIL", "SKIP", "미실행"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.column_dimensions["A"].width = 36
    for col_letter in "BCDEF":
        ws.column_dimensions[col_letter].width = 12

    row = 4
    for sheet_name, count in sorted(by_sheet.items()):
        # openpyxl 시트명 31자 제한
        safe = sheet_name[:31]
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=count)
        # COUNTIF 수식 — 결과 컬럼은 I
        ws.cell(row=row, column=3, value=f'=COUNTIF(\'{safe}\'!I:I,"PASS")')
        ws.cell(row=row, column=4, value=f'=COUNTIF(\'{safe}\'!I:I,"FAIL")')
        ws.cell(row=row, column=5, value=f'=COUNTIF(\'{safe}\'!I:I,"SKIP")')
        ws.cell(row=row, column=6, value=f'=B{row}-C{row}-D{row}-E{row}')
        row += 1

    # 합계 행
    total_row = row
    ws.cell(row=total_row, column=1, value="합계").font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{row-1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C4:C{row-1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{row-1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F4:F{row-1})")

    # 위험도 / 자동화 분포
    by_risk: dict[str, int] = {}
    by_auto: dict[str, int] = {}
    for c in cases:
        by_risk[c.risk] = by_risk.get(c.risk, 0) + 1
        by_auto[c.auto] = by_auto.get(c.auto, 0) + 1

    start_row = total_row + 3
    ws.cell(row=start_row, column=1, value="위험도 분포").font = Font(name="맑은 고딕", size=11, bold=True)
    for i, (k, v) in enumerate(sorted(by_risk.items()), start=1):
        ws.cell(row=start_row + i, column=1, value=k)
        ws.cell(row=start_row + i, column=2, value=v)

    start_row2 = start_row + len(by_risk) + 3
    ws.cell(row=start_row2, column=1, value="자동화 분포").font = Font(name="맑은 고딕", size=11, bold=True)
    for i, (k, v) in enumerate(sorted(by_auto.items()), start=1):
        ws.cell(row=start_row2 + i, column=1, value=k)
        ws.cell(row=start_row2 + i, column=2, value=v)


def main() -> None:
    cases = build_all_cases()
    print(f"[build_excel] cases: {len(cases)}")

    wb = Workbook()
    # 기본 시트 삭제 후 재구성
    default = wb.active
    wb.remove(default)

    # 화면별 시트로 분리
    by_sheet: dict[str, list] = {}
    for c in cases:
        by_sheet.setdefault(c.sheet, []).append(c)

    sheet_names = sorted(by_sheet.keys())

    # Cover / Summary 먼저
    _build_cover(wb, total_cases=len(cases), sheets=sheet_names)
    _build_summary(wb, cases)

    # 화면별 시트
    for sheet_name in sheet_names:
        # 시트명 31자 제한 + 특수문자 제거
        safe = sheet_name[:31].replace("/", "_").replace("\\", "_").replace("?", "")
        ws = wb.create_sheet(safe)
        _header(ws)
        for i, c in enumerate(by_sheet[sheet_name], start=2):
            _write_case(ws, i, c)
        _apply_conditional_formatting(ws, row_count=len(by_sheet[sheet_name]) + 1)
        ws.freeze_panes = "A2"
        # 행 높이 일괄
        for r in range(2, len(by_sheet[sheet_name]) + 2):
            ws.row_dimensions[r].height = 36

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"[build_excel] saved: {OUTPUT}")
    print(f"[build_excel] sheets: {len(sheet_names)} (+ Cover, Summary)")


if __name__ == "__main__":
    main()
