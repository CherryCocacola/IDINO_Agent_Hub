#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 원본 데이터 구조 분석기

기능:
1. Excel 파일의 모든 시트(탭) 목록 추출
2. 각 시트별 컬럼명, 데이터 타입, 행 수 분석
3. DB 스키마와 컬럼 매핑 검증
4. 불일치 항목 리포트 생성
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field
from datetime import datetime
import json

# openpyxl for Excel reading
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

# Database connection
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    print("psycopg2 설치 필요: pip install psycopg2-binary")
    sys.exit(1)

from migration_config import (
    DB_CONFIG, SCHEMA, TABLE_LEVELS,
    get_all_tables_in_order, TARGET_ROW_COUNTS
)

# ============================================
# Configuration
# ============================================

EXCEL_FILE = Path(r"E:\workspace\idino_career\user_mig\idino_table_columns.xlsx")
SHEET_PREFIX = "○tb_"  # Excel 시트 명명 규칙
OUTPUT_DIR = Path(r"E:\workspace\idino_career\database\analysis_reports")


@dataclass
class SheetInfo:
    """Excel 시트 정보"""
    sheet_name: str
    table_name: str  # "○tb_college" → "tb_college"
    columns: List[str] = field(default_factory=list)
    row_count: int = 0
    sample_data: List[Dict] = field(default_factory=list)
    column_types: Dict[str, str] = field(default_factory=dict)


@dataclass
class ColumnMapping:
    """컬럼 매핑 정보"""
    excel_column: str
    db_column: Optional[str]
    excel_type: str
    db_type: Optional[str]
    match_status: str  # "exact", "case_diff", "missing_in_db", "missing_in_excel"


@dataclass
class AnalysisReport:
    """분석 리포트"""
    excel_sheets: Dict[str, SheetInfo] = field(default_factory=dict)
    db_tables: Dict[str, List[str]] = field(default_factory=dict)
    column_mappings: Dict[str, List[ColumnMapping]] = field(default_factory=dict)
    row_count_comparison: Dict[str, Tuple[int, int]] = field(default_factory=dict)  # (excel, db)
    issues: List[str] = field(default_factory=list)


def get_db_connection():
    """DB 연결 생성"""
    return psycopg2.connect(**DB_CONFIG)


def extract_table_name(sheet_name: str) -> str:
    """시트명에서 테이블명 추출: '○tb_college' → 'tb_college'"""
    if sheet_name.startswith(SHEET_PREFIX):
        return sheet_name[1:]  # "○" 제거
    elif sheet_name.startswith("tb_"):
        return sheet_name
    return sheet_name


def infer_column_type(values: List[Any]) -> str:
    """샘플 값들로부터 데이터 타입 추론"""
    non_null_values = [v for v in values if v is not None]
    if not non_null_values:
        return "unknown"

    # 타입별 카운트
    type_counts = {"int": 0, "float": 0, "str": 0, "date": 0, "bool": 0}

    for v in non_null_values[:100]:  # 최대 100개 샘플
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, int):
            type_counts["int"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        elif isinstance(v, datetime):
            type_counts["date"] += 1
        elif isinstance(v, str):
            # 문자열이지만 숫자일 수 있음
            try:
                int(v)
                type_counts["int"] += 1
            except ValueError:
                try:
                    float(v)
                    type_counts["float"] += 1
                except ValueError:
                    type_counts["str"] += 1
        else:
            type_counts["str"] += 1

    # 가장 많은 타입 반환
    return max(type_counts, key=type_counts.get)


def analyze_excel_file(excel_path: Path, max_rows_sample: int = 10) -> Dict[str, SheetInfo]:
    """Excel 파일의 모든 시트 분석"""
    print(f"\n{'='*60}")
    print(f"Excel 파일 분석 시작: {excel_path}")
    print(f"파일 크기: {excel_path.stat().st_size / (1024*1024):.1f} MB")
    print(f"{'='*60}")

    # read_only 모드로 대용량 파일 처리
    print("Excel 파일 로드 중... (대용량 파일이므로 시간이 걸릴 수 있습니다)")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    sheets_info = {}
    all_sheet_names = wb.sheetnames

    # tb_ 패턴 시트만 필터링
    target_sheets = [s for s in all_sheet_names if SHEET_PREFIX in s or s.startswith("tb_")]

    print(f"\n전체 시트 수: {len(all_sheet_names)}")
    print(f"테이블 시트 수 ({SHEET_PREFIX}* 패턴): {len(target_sheets)}")

    for idx, sheet_name in enumerate(target_sheets, 1):
        print(f"\n[{idx}/{len(target_sheets)}] 시트 분석 중: {sheet_name}")

        ws = wb[sheet_name]
        table_name = extract_table_name(sheet_name)

        # 첫 번째 행은 헤더
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            print(f"  [WARN] 빈 시트")
            continue

        header_row = rows[0]
        columns = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(header_row)]

        # 데이터 행 (헤더 제외)
        data_rows = rows[1:]
        row_count = len(data_rows)

        # 샘플 데이터 추출
        sample_data = []
        for row in data_rows[:max_rows_sample]:
            row_dict = {columns[i]: row[i] for i in range(min(len(columns), len(row)))}
            sample_data.append(row_dict)

        # 컬럼 타입 추론
        column_types = {}
        for col_idx, col_name in enumerate(columns):
            col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows[:100]]
            column_types[col_name] = infer_column_type(col_values)

        sheet_info = SheetInfo(
            sheet_name=sheet_name,
            table_name=table_name,
            columns=columns,
            row_count=row_count,
            sample_data=sample_data,
            column_types=column_types
        )

        sheets_info[table_name] = sheet_info
        print(f"  [OK] 컬럼: {len(columns)}개, 행: {row_count:,}개")

    wb.close()
    return sheets_info


def get_db_schema() -> Dict[str, List[Tuple[str, str]]]:
    """DB 스키마에서 테이블별 컬럼 정보 조회"""
    print(f"\n{'='*60}")
    print(f"DB 스키마 조회: {SCHEMA}")
    print(f"{'='*60}")

    conn = get_db_connection()
    cur = conn.cursor()

    # 모든 테이블의 컬럼 정보 조회
    cur.execute("""
        SELECT
            table_name,
            column_name,
            data_type,
            character_maximum_length,
            is_nullable,
            column_default
        FROM information_schema.columns
        WHERE table_schema = %s
        ORDER BY table_name, ordinal_position
    """, (SCHEMA,))

    rows = cur.fetchall()

    db_tables = {}
    for row in rows:
        table_name, col_name, data_type, max_len, nullable, default = row
        if table_name not in db_tables:
            db_tables[table_name] = []

        # 타입 포맷팅
        if max_len:
            type_str = f"{data_type}({max_len})"
        else:
            type_str = data_type

        db_tables[table_name].append((col_name, type_str))

    cur.close()
    conn.close()

    print(f"DB 테이블 수: {len(db_tables)}")
    return db_tables


def get_db_row_counts() -> Dict[str, int]:
    """DB 테이블별 행 수 조회"""
    conn = get_db_connection()
    cur = conn.cursor()

    row_counts = {}
    tables = get_all_tables_in_order()

    for table in tables:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.{table}")
            count = cur.fetchone()[0]
            row_counts[table] = count
        except Exception as e:
            row_counts[table] = -1  # 테이블 없음 또는 오류

    cur.close()
    conn.close()

    return row_counts


def compare_columns(excel_info: SheetInfo, db_columns: List[Tuple[str, str]]) -> List[ColumnMapping]:
    """Excel 컬럼과 DB 컬럼 비교"""
    mappings = []

    excel_cols = {c.lower(): c for c in excel_info.columns}
    db_cols = {c[0].lower(): (c[0], c[1]) for c in db_columns}

    # Excel 컬럼 기준으로 비교
    for excel_col in excel_info.columns:
        excel_col_lower = excel_col.lower()
        excel_type = excel_info.column_types.get(excel_col, "unknown")

        if excel_col_lower in db_cols:
            db_col_name, db_type = db_cols[excel_col_lower]
            if excel_col == db_col_name:
                status = "exact"
            else:
                status = "case_diff"

            mappings.append(ColumnMapping(
                excel_column=excel_col,
                db_column=db_col_name,
                excel_type=excel_type,
                db_type=db_type,
                match_status=status
            ))
        else:
            mappings.append(ColumnMapping(
                excel_column=excel_col,
                db_column=None,
                excel_type=excel_type,
                db_type=None,
                match_status="missing_in_db"
            ))

    # DB에만 있는 컬럼
    for db_col_lower, (db_col_name, db_type) in db_cols.items():
        if db_col_lower not in excel_cols:
            mappings.append(ColumnMapping(
                excel_column=None,
                db_column=db_col_name,
                excel_type=None,
                db_type=db_type,
                match_status="missing_in_excel"
            ))

    return mappings


def generate_report(
    excel_sheets: Dict[str, SheetInfo],
    db_tables: Dict[str, List[Tuple[str, str]]],
    db_row_counts: Dict[str, int]
) -> AnalysisReport:
    """전체 분석 리포트 생성"""
    report = AnalysisReport()
    report.excel_sheets = excel_sheets
    report.db_tables = {k: [c[0] for c in v] for k, v in db_tables.items()}

    print(f"\n{'='*60}")
    print("컬럼 매핑 및 행 수 비교")
    print(f"{'='*60}")

    # 모든 예상 테이블 목록
    expected_tables = set(get_all_tables_in_order())

    for table_name in sorted(expected_tables | set(excel_sheets.keys())):
        excel_info = excel_sheets.get(table_name)
        db_cols = db_tables.get(table_name, [])

        # 행 수 비교
        excel_rows = excel_info.row_count if excel_info else 0
        db_rows = db_row_counts.get(table_name, 0)
        report.row_count_comparison[table_name] = (excel_rows, db_rows)

        # 컬럼 매핑
        if excel_info and db_cols:
            mappings = compare_columns(excel_info, db_cols)
            report.column_mappings[table_name] = mappings

            # 불일치 이슈 기록
            missing_in_db = [m for m in mappings if m.match_status == "missing_in_db"]
            missing_in_excel = [m for m in mappings if m.match_status == "missing_in_excel"]

            if missing_in_db:
                report.issues.append(
                    f"{table_name}: Excel에만 있는 컬럼 {len(missing_in_db)}개 - "
                    f"{[m.excel_column for m in missing_in_db]}"
                )
            if missing_in_excel:
                report.issues.append(
                    f"{table_name}: DB에만 있는 컬럼 {len(missing_in_excel)}개 - "
                    f"{[m.db_column for m in missing_in_excel]}"
                )

        elif not excel_info and table_name in expected_tables:
            report.issues.append(f"{table_name}: Excel 시트 없음")

        elif excel_info and not db_cols:
            report.issues.append(f"{table_name}: DB 테이블 없음")

    return report


def print_report(report: AnalysisReport):
    """리포트 출력"""
    print(f"\n{'='*60}")
    print("[REPORT] 분석 결과 요약")
    print(f"{'='*60}")

    # 행 수 비교 테이블
    print("\n### 테이블별 행 수 비교")
    print(f"{'테이블':<30} {'Excel':>10} {'DB':>10} {'차이':>10} {'상태':>8}")
    print("-" * 72)

    total_excel = 0
    total_db = 0

    for table_name, (excel_rows, db_rows) in sorted(report.row_count_comparison.items()):
        diff = excel_rows - db_rows
        total_excel += excel_rows
        total_db += max(0, db_rows)

        if excel_rows > 0 and db_rows >= 0:
            if diff > 0:
                status = "[!] 부족"
            elif diff < 0:
                status = "[!] 초과"
            else:
                status = "[OK]"
        elif excel_rows == 0:
            status = "[?] 없음"
        else:
            status = "[X] 오류"

        print(f"{table_name:<30} {excel_rows:>10,} {db_rows:>10,} {diff:>+10,} {status:>8}")

    print("-" * 72)
    print(f"{'합계':<30} {total_excel:>10,} {total_db:>10,} {total_excel - total_db:>+10,}")

    # 주요 이슈
    if report.issues:
        print(f"\n### [ISSUES] 발견된 이슈 ({len(report.issues)}건)")
        for issue in report.issues[:20]:  # 최대 20건
            print(f"  - {issue}")
        if len(report.issues) > 20:
            print(f"  ... 외 {len(report.issues) - 20}건")

    # Excel 시트 목록
    print(f"\n### Excel 시트 목록 ({len(report.excel_sheets)}개)")
    for table_name, sheet_info in sorted(report.excel_sheets.items()):
        print(f"  - {sheet_info.sheet_name}: {sheet_info.row_count:,}행, {len(sheet_info.columns)}컬럼")


def save_report_json(report: AnalysisReport, output_path: Path):
    """리포트를 JSON으로 저장"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # dataclass를 dict로 변환
    report_dict = {
        "generated_at": datetime.now().isoformat(),
        "excel_sheets": {
            k: {
                "sheet_name": v.sheet_name,
                "table_name": v.table_name,
                "columns": v.columns,
                "row_count": v.row_count,
                "column_types": v.column_types
            }
            for k, v in report.excel_sheets.items()
        },
        "db_tables": report.db_tables,
        "row_count_comparison": {
            k: {"excel": v[0], "db": v[1], "diff": v[0] - v[1]}
            for k, v in report.row_count_comparison.items()
        },
        "column_mappings": {
            k: [
                {
                    "excel_column": m.excel_column,
                    "db_column": m.db_column,
                    "excel_type": m.excel_type,
                    "db_type": m.db_type,
                    "match_status": m.match_status
                }
                for m in v
            ]
            for k, v in report.column_mappings.items()
        },
        "issues": report.issues
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report_dict, f, ensure_ascii=False, indent=2)

    print(f"\n[SAVED] JSON 리포트 저장: {output_path}")


def save_column_mapping_csv(report: AnalysisReport, output_path: Path):
    """컬럼 매핑 테이블을 CSV로 저장"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w', encoding='utf-8-sig') as f:
        f.write("table_name,excel_column,db_column,excel_type,db_type,match_status\n")

        for table_name, mappings in sorted(report.column_mappings.items()):
            for m in mappings:
                f.write(f"{table_name},{m.excel_column or ''},{m.db_column or ''},"
                       f"{m.excel_type or ''},{m.db_type or ''},{m.match_status}\n")

    print(f"[SAVED] 컬럼 매핑 CSV 저장: {output_path}")


def main():
    """메인 실행"""
    print("=" * 60)
    print("IDINO Career - Excel 원본 데이터 구조 분석기")
    print("=" * 60)

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel 파일 없음: {EXCEL_FILE}")
        sys.exit(1)

    # 1. Excel 분석
    excel_sheets = analyze_excel_file(EXCEL_FILE)

    # 2. DB 스키마 조회
    db_tables = get_db_schema()
    db_row_counts = get_db_row_counts()

    # 3. 리포트 생성
    report = generate_report(excel_sheets, db_tables, db_row_counts)

    # 4. 결과 출력
    print_report(report)

    # 5. 파일 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    save_report_json(report, OUTPUT_DIR / f"analysis_report_{timestamp}.json")
    save_column_mapping_csv(report, OUTPUT_DIR / f"column_mappings_{timestamp}.csv")

    # 6. 요약 통계
    print(f"\n{'='*60}")
    print("[SUMMARY] 최종 요약")
    print(f"{'='*60}")
    print(f"Excel 시트 (테이블): {len(excel_sheets)}개")
    print(f"DB 테이블: {len(db_tables)}개")
    print(f"발견된 이슈: {len(report.issues)}건")

    # 마이그레이션 대상 요약
    migration_targets = []
    for table_name, (excel_rows, db_rows) in report.row_count_comparison.items():
        if excel_rows > 0 and excel_rows != db_rows:
            migration_targets.append((table_name, excel_rows, db_rows))

    if migration_targets:
        print(f"\n### 마이그레이션 필요 테이블 ({len(migration_targets)}개)")
        for table, excel, db in sorted(migration_targets, key=lambda x: x[1], reverse=True)[:10]:
            print(f"  - {table}: Excel {excel:,}건 vs DB {db:,}건")

    return report


if __name__ == "__main__":
    main()
