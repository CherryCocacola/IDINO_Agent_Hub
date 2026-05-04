"""
Excel → SQL 변환 스크립트
idino_table_columns.xlsx 에서 tb_competency, tb_student_competency 데이터를 읽어
52_update_competency_from_excel.sql 파일을 자동 생성합니다.

Usage:
    pip install openpyxl
    python database/52_import_excel_competency.py
"""
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 패키지가 필요합니다. 설치: pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
EXCEL_PATH = SCRIPT_DIR.parent / "user_mig" / "idino_table_columns.xlsx"
OUTPUT_PATH = SCRIPT_DIR / "52_update_competency_from_excel.sql"

# Active competency codes (use_fg='Y') - codes 1-5 (COMP001~COMP006 or numeric 1~5)
# Codes 6-14 are inactive and their student_competency records should be deleted
ACTIVE_COMPETENCY_CODES = set()  # Will be populated from Excel
INACTIVE_COMPETENCY_RANGE = range(6, 15)  # 6~14 are inactive


def find_sheet(wb, candidates):
    """Find a sheet by trying multiple name candidates."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # Try partial match
    for sheet_name in wb.sheetnames:
        for candidate in candidates:
            if candidate.lower() in sheet_name.lower():
                return wb[sheet_name]
    return None


def read_competency_sheet(wb):
    """Read tb_competency sheet to identify active/inactive codes."""
    sheet = find_sheet(wb, ['tb_competency', 'competency', 'TB_COMPETENCY'])
    if not sheet:
        print(f"Warning: tb_competency 시트를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
        return {}

    # Find header row
    headers = {}
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ''
            if 'competency_cd' in val or 'competency_nm' in val or 'use_fg' in val:
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("Warning: tb_competency 시트에서 헤더를 찾을 수 없습니다.")
        return {}

    # Parse headers
    for cell in sheet[header_row]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column - 1

    # Read competency data
    competencies = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        values = list(row)

        cd_idx = headers.get('competency_cd', 0)
        nm_idx = headers.get('competency_nm', 1)
        use_fg_idx = headers.get('use_fg')

        cd = str(values[cd_idx]).strip() if cd_idx < len(values) and values[cd_idx] else None
        nm = str(values[nm_idx]).strip() if nm_idx is not None and nm_idx < len(values) and values[nm_idx] else ''
        use_fg = str(values[use_fg_idx]).strip().upper() if use_fg_idx is not None and use_fg_idx < len(values) and values[use_fg_idx] else 'Y'

        if cd:
            competencies[cd] = {'nm': nm, 'use_fg': use_fg}
            if use_fg == 'Y':
                ACTIVE_COMPETENCY_CODES.add(cd)

    print(f"읽은 역량 정의: {len(competencies)}개 (활성: {len(ACTIVE_COMPETENCY_CODES)}개)")
    return competencies


def read_student_competency_sheet(wb):
    """Read tb_student_competency sheet to get student competency scores."""
    sheet = find_sheet(wb, ['tb_student_competency', 'student_competency', 'TB_STUDENT_COMPETENCY'])
    if not sheet:
        print(f"Warning: tb_student_competency 시트를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
        return []

    # Find header row
    headers = {}
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ''
            if 'student_id' in val or 'competency_cd' in val or 'current_score' in val:
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("Warning: tb_student_competency 시트에서 헤더를 찾을 수 없습니다.")
        return []

    # Parse headers
    for cell in sheet[header_row]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column - 1

    required = ['student_id', 'competency_cd', 'current_score']
    for col in required:
        if col not in headers:
            print(f"Warning: 필수 컬럼 '{col}'을 찾을 수 없습니다. 존재하는 컬럼: {list(headers.keys())}")
            return []

    # Read data
    records = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        values = list(row)

        student_id = str(values[headers['student_id']]).strip() if values[headers['student_id']] else None
        competency_cd = str(values[headers['competency_cd']]).strip() if values[headers['competency_cd']] else None
        current_score_raw = values[headers['current_score']]

        if not student_id or not competency_cd:
            continue

        try:
            current_score = float(current_score_raw) if current_score_raw is not None else 0.0
        except (ValueError, TypeError):
            current_score = 0.0

        # Also read target_score and status if available
        target_score = None
        if 'target_score' in headers:
            try:
                ts = values[headers['target_score']]
                target_score = float(ts) if ts is not None else None
            except (ValueError, TypeError):
                pass

        status = None
        if 'status' in headers:
            s = values[headers['status']]
            status = str(s).strip() if s else None

        records.append({
            'student_id': student_id,
            'competency_cd': competency_cd,
            'current_score': current_score,
            'target_score': target_score,
            'status': status,
        })

    print(f"읽은 학생역량 레코드: {len(records)}개")
    return records


def generate_sql(competencies, student_records):
    """Generate SQL file from extracted data."""
    lines = []
    lines.append("-- =============================================================")
    lines.append("-- 52_update_competency_from_excel.sql")
    lines.append("-- Excel(idino_table_columns.xlsx)에서 추출한 역량 점수 UPDATE")
    lines.append("-- 자동 생성 스크립트: 52_import_excel_competency.py")
    lines.append("-- =============================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # Part 1: Update competency use_fg based on Excel
    if competencies:
        lines.append("-- Part 1: 역량 정의 use_fg 업데이트")
        for cd, info in competencies.items():
            use_fg = info['use_fg']
            lines.append(f"UPDATE tb_competency SET use_fg = '{use_fg}' WHERE competency_cd = '{cd}';")
        lines.append("")

    # Part 2: Delete inactive competency records
    inactive_cds = [cd for cd, info in competencies.items() if info['use_fg'] != 'Y']
    if inactive_cds:
        lines.append("-- Part 2: 비활성 역량(use_fg='N')의 student_competency 레코드 삭제")
        cd_list = "', '".join(inactive_cds)
        lines.append(f"DELETE FROM tb_student_competency WHERE competency_cd IN ('{cd_list}');")
        lines.append("")

    # Part 3: Update student competency scores
    if student_records:
        lines.append("-- Part 3: 학생별 역량 점수 UPDATE (Excel 원본 값)")

        # Group by student
        student_groups = {}
        for rec in student_records:
            sid = rec['student_id']
            if sid not in student_groups:
                student_groups[sid] = []
            student_groups[sid].append(rec)

        lines.append(f"-- 총 학생 수: {len(student_groups)}명")
        lines.append("")

        for student_id in sorted(student_groups.keys()):
            recs = student_groups[student_id]
            lines.append(f"-- Student: {student_id}")
            for rec in recs:
                cd = rec['competency_cd']
                score = rec['current_score']

                # Skip inactive competencies
                if cd in [c for c, i in competencies.items() if i['use_fg'] != 'Y']:
                    continue

                # Build SET clause
                set_parts = [f"current_score = {score}"]
                if rec['target_score'] is not None:
                    set_parts.append(f"target_score = {rec['target_score']}")
                if rec['status']:
                    set_parts.append(f"status = '{rec['status']}'")
                set_parts.append(f"gap_score = {score} - COALESCE(target_score, 85)")
                set_parts.append("upd_dt = CURRENT_TIMESTAMP")
                set_parts.append("upd_user_id = 'EXCEL_IMPORT'")

                set_clause = ", ".join(set_parts)
                lines.append(
                    f"UPDATE tb_student_competency SET {set_clause} "
                    f"WHERE student_id = '{student_id}' AND competency_cd = '{cd}';"
                )

                # Insert if not exists
                lines.append(
                    f"INSERT INTO tb_student_competency (student_id, competency_cd, current_score, target_score, gap_score, status, ins_user_id, ins_dt) "
                    f"SELECT '{student_id}', '{cd}', {score}, 85, {score} - 85, "
                    f"CASE WHEN {score} >= 400 THEN 'excellent' WHEN {score} >= 300 THEN 'good' "
                    f"WHEN {score} >= 200 THEN 'average' ELSE 'improve' END, "
                    f"'EXCEL_IMPORT', CURRENT_TIMESTAMP "
                    f"WHERE NOT EXISTS (SELECT 1 FROM tb_student_competency WHERE student_id = '{student_id}' AND competency_cd = '{cd}');"
                )
            lines.append("")

    lines.append("COMMIT;")
    lines.append("")
    lines.append(f"-- Total records processed: {len(student_records)}")

    return "\n".join(lines)


def main():
    if not EXCEL_PATH.exists():
        print(f"Error: Excel 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Excel 파일 로딩: {EXCEL_PATH}")
    print("(대용량 파일의 경우 시간이 소요될 수 있습니다...)")

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"시트 목록: {wb.sheetnames}")

    # Read competency definitions
    competencies = read_competency_sheet(wb)

    # Read student competency scores
    student_records = read_student_competency_sheet(wb)

    wb.close()

    if not student_records:
        print("Warning: 학생 역량 데이터를 읽지 못했습니다.")
        print("시트명이나 컬럼명을 확인해주세요.")
        # Still generate SQL for competency use_fg updates
        if not competencies:
            print("역량 데이터도 없습니다. SQL 생성을 건너뜁니다.")
            return

    # Generate SQL
    sql = generate_sql(competencies, student_records)

    # Write output
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f"\nSQL 파일 생성 완료: {OUTPUT_PATH}")
    print(f"  - 역량 정의: {len(competencies)}개")
    print(f"  - 학생 레코드: {len(student_records)}개")
    if competencies:
        active = sum(1 for c in competencies.values() if c['use_fg'] == 'Y')
        inactive = len(competencies) - active
        print(f"  - 활성 역량: {active}개, 비활성 역량: {inactive}개")


if __name__ == "__main__":
    main()
