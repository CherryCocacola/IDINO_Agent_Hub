#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 원본 데이터 → DB 마이그레이션 스크립트

기능:
1. Excel 시트 데이터 읽기 (openpyxl read_only 모드)
2. 컬럼 매핑 (Excel 컬럼명 → DB 컬럼명)
3. FK 의존성 순서대로 INSERT
4. 무결성 검증 후 COMMIT/ROLLBACK
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Generator
from dataclasses import dataclass
from datetime import datetime, date
import json

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

try:
    import psycopg2
    from psycopg2.extras import execute_values, RealDictCursor
except ImportError:
    print("psycopg2 설치 필요: pip install psycopg2-binary")
    sys.exit(1)

from migration_config import DB_CONFIG, SCHEMA

# ============================================
# Configuration
# ============================================

EXCEL_FILE = Path(r"E:\workspace\idino_career\user_mig\idino_table_columns.xlsx")
SHEET_PREFIX = "○"
BATCH_SIZE = 1000  # 배치 INSERT 크기

# ============================================
# 테이블 매핑 설정 (Excel 시트명 → DB 테이블명)
# ============================================

TABLE_NAME_MAPPING = {
    "tb_student2": "tb_student",  # Excel의 tb_student2 → DB의 tb_student
}

# ============================================
# FK 의존성 순서 (Excel에 있는 테이블만)
# ============================================

MIGRATION_ORDER = [
    # Level 1-2: 기본 마스터 (Excel에 있는 것)
    ("tb_college", "○tb_college"),
    ("tb_competency", "○tb_competency"),
    ("tb_program", "○tb_program"),

    # Level 3: department, course
    ("tb_department", "○tb_department"),
    ("tb_course", "○tb_course"),

    # Level 4: professor, student
    ("tb_professor", "○tb_professor"),
    ("tb_student", "○tb_student2"),  # tb_student2 → tb_student로 매핑

    # Level 5: course_offering, professor_course, user
    ("tb_course_offering", "○tb_course_offering"),
    ("tb_professor_course", "○tb_professor_course"),
    ("tb_user", "○tb_user"),

    # Level 6: enrollment, competency mappings, activities
    ("tb_enrollment", "○tb_enrollment"),
    ("tb_course_competency_map", "○tb_course_competency_map"),
    ("tb_student_competency", "○tb_student_competency"),
    ("tb_activity", "○tb_activity"),
    ("tb_achievement", "tb_achievement"),

    # Level 7: grades, summaries
    ("tb_grade", "○tb_grade"),
    ("tb_grade_summary", "○tb_grade_summary"),
    ("tb_cumulative_summary", "○tb_cumulative_summary"),
]

# ============================================
# 컬럼 매핑 (col_X 같은 빈 컬럼 제외)
# ============================================

# 테이블별 컬럼명 매핑 (Excel 컬럼명 → DB 컬럼명)
COLUMN_MAPPINGS = {
    "tb_enrollment": {
        "offering_id": "course_offering_id",
        "status": "status_cd",
    },
    # 다른 테이블은 자동 매핑
}

# 제외할 컬럼 패턴 (정규식 아님, startswith로 체크)
EXCLUDE_COLUMN_PATTERNS = ["col_"]


def get_db_connection():
    """DB 연결 생성"""
    return psycopg2.connect(**DB_CONFIG)


def get_db_columns(table_name: str) -> List[str]:
    """DB 테이블의 컬럼 목록 조회"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
    """, (SCHEMA, table_name))
    columns = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()
    return columns


def should_exclude_column(col_name: str) -> bool:
    """컬럼 제외 여부 확인"""
    if not col_name:
        return True
    for pattern in EXCLUDE_COLUMN_PATTERNS:
        if col_name.startswith(pattern):
            return True
    # 쉼표가 포함된 컬럼명도 제외 (Excel 오류)
    if "," in col_name:
        return True
    return False


def convert_value(value: Any, db_type: str = None) -> Any:
    """값 변환 (Excel → DB 타입)"""
    if value is None:
        return None

    # datetime 객체 처리
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value

    # 문자열 처리
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value.lower() == "null":
            return None
        return value

    # 숫자 처리
    if isinstance(value, (int, float)):
        return value

    return value


def read_excel_sheet(workbook, sheet_name: str, db_table: str = None) -> Generator[Dict[str, Any], None, None]:
    """Excel 시트에서 데이터 읽기 (Generator로 메모리 효율화)"""
    if sheet_name not in workbook.sheetnames:
        print(f"  [WARN] 시트 없음: {sheet_name}")
        return

    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return

    # 헤더 (첫 번째 행)
    header = rows[0]
    columns = [str(c).strip() if c else f"_empty_{i}" for i, c in enumerate(header)]

    # 컬럼 매핑 가져오기
    col_mapping = COLUMN_MAPPINGS.get(db_table, {})

    # 데이터 행
    for row in rows[1:]:
        row_dict = {}
        for i, col_name in enumerate(columns):
            if should_exclude_column(col_name):
                continue
            value = row[i] if i < len(row) else None

            # 컬럼명 매핑 적용
            mapped_col_name = col_mapping.get(col_name, col_name)
            row_dict[mapped_col_name] = convert_value(value)

        # 빈 행 스킵
        if all(v is None for v in row_dict.values()):
            continue

        yield row_dict


def get_insert_columns(excel_columns: List[str], db_columns: List[str]) -> List[str]:
    """Excel과 DB에서 공통으로 사용할 컬럼 결정"""
    # Excel 컬럼 중 DB에 존재하는 것만 선택
    excel_cols_lower = {c.lower(): c for c in excel_columns if not should_exclude_column(c)}
    db_cols_lower = {c.lower(): c for c in db_columns}

    common_cols = []
    for excel_col_lower, excel_col in excel_cols_lower.items():
        if excel_col_lower in db_cols_lower:
            common_cols.append(db_cols_lower[excel_col_lower])

    return common_cols


def truncate_table(conn, table_name: str):
    """테이블 데이터 삭제 (CASCADE)"""
    cur = conn.cursor()
    try:
        cur.execute(f"TRUNCATE TABLE {SCHEMA}.{table_name} CASCADE")
        conn.commit()
        print(f"  [TRUNCATE] {table_name} 완료")
    except Exception as e:
        conn.rollback()
        print(f"  [ERROR] TRUNCATE {table_name} 실패: {e}")
        raise
    finally:
        cur.close()


def insert_batch(conn, table_name: str, columns: List[str], rows: List[Dict]) -> int:
    """배치 INSERT 실행"""
    if not rows:
        return 0

    cur = conn.cursor()

    # INSERT 쿼리 생성
    col_str = ", ".join(columns)
    placeholders = ", ".join(["%s"] * len(columns))

    query = f"INSERT INTO {SCHEMA}.{table_name} ({col_str}) VALUES ({placeholders})"

    # 값 준비
    values = []
    for row in rows:
        row_values = tuple(row.get(col.lower()) or row.get(col) for col in columns)
        values.append(row_values)

    try:
        cur.executemany(query, values)
        conn.commit()
        return len(values)
    except Exception as e:
        conn.rollback()
        # 첫 번째 실패 행 출력
        print(f"  [ERROR] INSERT 실패: {e}")
        if values:
            print(f"  [DEBUG] 첫 번째 행: {dict(zip(columns, values[0]))}")
        raise
    finally:
        cur.close()


def migrate_table(workbook, db_table: str, excel_sheet: str, conn) -> Tuple[int, int]:
    """단일 테이블 마이그레이션"""
    print(f"\n[MIGRATE] {db_table} <- {excel_sheet}")

    # DB 컬럼 조회
    db_columns = get_db_columns(db_table)
    if not db_columns:
        print(f"  [ERROR] DB 테이블 없음: {db_table}")
        return 0, 0

    # Excel 시트 헤더 확인
    if excel_sheet not in workbook.sheetnames:
        print(f"  [SKIP] Excel 시트 없음: {excel_sheet}")
        return 0, 0

    ws = workbook[excel_sheet]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        print(f"  [SKIP] 빈 시트: {excel_sheet}")
        return 0, 0

    excel_columns = [str(c).strip() if c else f"_empty_{i}" for i, c in enumerate(first_row)]

    # 컬럼 매핑 적용
    col_mapping = COLUMN_MAPPINGS.get(db_table, {})
    mapped_excel_columns = [col_mapping.get(c, c) for c in excel_columns]

    # 공통 컬럼 결정
    insert_columns = get_insert_columns(mapped_excel_columns, db_columns)
    if not insert_columns:
        print(f"  [ERROR] 공통 컬럼 없음")
        return 0, 0

    print(f"  DB 컬럼: {len(db_columns)}개")
    print(f"  Excel 컬럼: {len([c for c in excel_columns if not should_exclude_column(c)])}개")
    print(f"  매핑 컬럼: {len(insert_columns)}개")

    # 기존 데이터 삭제
    truncate_table(conn, db_table)

    # 배치 INSERT
    batch = []
    total_inserted = 0
    total_rows = 0

    for row_dict in read_excel_sheet(workbook, excel_sheet, db_table):
        total_rows += 1

        # 컬럼명을 소문자로 정규화
        normalized_row = {k.lower(): v for k, v in row_dict.items()}
        batch.append(normalized_row)

        if len(batch) >= BATCH_SIZE:
            inserted = insert_batch(conn, db_table, insert_columns, batch)
            total_inserted += inserted
            print(f"  진행: {total_inserted:,}/{total_rows:,} ({100*total_inserted//max(1,total_rows)}%)", end="\r")
            batch = []

    # 나머지 배치 처리
    if batch:
        inserted = insert_batch(conn, db_table, insert_columns, batch)
        total_inserted += inserted

    print(f"  [DONE] {total_inserted:,}건 삽입 완료                    ")
    return total_inserted, total_rows


def verify_fk_integrity(conn) -> List[str]:
    """FK 무결성 검증"""
    print("\n" + "=" * 60)
    print("[VERIFY] FK 무결성 검증")
    print("=" * 60)

    issues = []

    # 검증할 FK 관계
    fk_checks = [
        ("tb_department", "college_cd", "tb_college", "college_cd"),
        ("tb_course", "department_cd", "tb_department", "department_cd"),
        ("tb_student", "department_cd", "tb_department", "department_cd"),
        ("tb_student", "university_cd", "tb_university", "university_cd"),
        ("tb_professor", "department_cd", "tb_department", "department_cd"),
        ("tb_course_offering", "course_cd", "tb_course", "course_cd"),
        ("tb_enrollment", "student_id", "tb_student", "student_id"),
        ("tb_grade", "enrollment_id", "tb_enrollment", "enrollment_id"),
    ]

    cur = conn.cursor()

    for child_table, child_col, parent_table, parent_col in fk_checks:
        try:
            query = f"""
                SELECT COUNT(*) FROM {SCHEMA}.{child_table} c
                LEFT JOIN {SCHEMA}.{parent_table} p ON c.{child_col} = p.{parent_col}
                WHERE c.{child_col} IS NOT NULL AND p.{parent_col} IS NULL
            """
            cur.execute(query)
            orphan_count = cur.fetchone()[0]

            if orphan_count > 0:
                issue = f"{child_table}.{child_col} -> {parent_table}.{parent_col}: {orphan_count}건 참조 누락"
                issues.append(issue)
                print(f"  [WARN] {issue}")
            else:
                print(f"  [OK] {child_table}.{child_col} -> {parent_table}.{parent_col}")
        except Exception as e:
            print(f"  [SKIP] {child_table}.{child_col}: {e}")

    cur.close()
    return issues


def print_row_counts(conn):
    """테이블별 행 수 출력"""
    print("\n" + "=" * 60)
    print("[RESULT] 마이그레이션 결과")
    print("=" * 60)

    cur = conn.cursor()

    print(f"{'테이블':<30} {'행 수':>12}")
    print("-" * 45)

    total = 0
    for db_table, _ in MIGRATION_ORDER:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.{db_table}")
            count = cur.fetchone()[0]
            total += count
            print(f"{db_table:<30} {count:>12,}")
        except Exception as e:
            print(f"{db_table:<30} {'ERROR':>12}")

    print("-" * 45)
    print(f"{'합계':<30} {total:>12,}")

    cur.close()


def main():
    """메인 실행"""
    print("=" * 60)
    print("IDINO Career - Excel 원본 데이터 마이그레이션")
    print("=" * 60)

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel 파일 없음: {EXCEL_FILE}")
        sys.exit(1)

    print(f"Excel 파일: {EXCEL_FILE}")
    print(f"파일 크기: {EXCEL_FILE.stat().st_size / (1024*1024):.1f} MB")
    print(f"대상 스키마: {SCHEMA}")
    print(f"마이그레이션 테이블 수: {len(MIGRATION_ORDER)}개")

    # Excel 파일 로드
    print("\nExcel 파일 로드 중... (시간이 걸릴 수 있습니다)")
    workbook = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    print(f"시트 수: {len(workbook.sheetnames)}")

    # DB 연결
    conn = get_db_connection()
    conn.autocommit = False

    try:
        # FK 제약조건 임시 비활성화
        cur = conn.cursor()
        cur.execute("SET session_replication_role = 'replica';")
        conn.commit()
        cur.close()
        print("\n[INFO] FK 제약조건 임시 비활성화")

        # 테이블별 마이그레이션
        results = {}
        for db_table, excel_sheet in MIGRATION_ORDER:
            try:
                inserted, total = migrate_table(workbook, db_table, excel_sheet, conn)
                results[db_table] = (inserted, total, "OK")
            except Exception as e:
                results[db_table] = (0, 0, f"ERROR: {e}")
                print(f"  [ERROR] {db_table} 실패: {e}")
                # 에러 발생해도 계속 진행

        # FK 제약조건 다시 활성화
        cur = conn.cursor()
        cur.execute("SET session_replication_role = 'origin';")
        conn.commit()
        cur.close()
        print("\n[INFO] FK 제약조건 다시 활성화")

        # 결과 출력
        print_row_counts(conn)

        # FK 무결성 검증
        issues = verify_fk_integrity(conn)

        if issues:
            print(f"\n[WARN] FK 무결성 이슈 {len(issues)}건 발견")
        else:
            print("\n[OK] FK 무결성 검증 통과")

    except Exception as e:
        conn.rollback()
        print(f"\n[FATAL] 마이그레이션 실패: {e}")
        raise
    finally:
        workbook.close()
        conn.close()

    print("\n" + "=" * 60)
    print("마이그레이션 완료")
    print("=" * 60)


if __name__ == "__main__":
    main()
