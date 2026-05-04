#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 원본 데이터 → DB 마이그레이션 스크립트 (V2 - 데이터 변환 포함)

개선 사항:
1. UUID 변환: 원본 ID → uuid5 (결정론적 변환)
2. DATE 변환: 20200522 → 2020-05-22
3. NOT NULL 기본값 설정
4. 중복 키 처리 (ON CONFLICT DO NOTHING)
"""

import os
import sys
import uuid
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Generator, Set
from dataclasses import dataclass
from datetime import datetime, date
import json
import re

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

try:
    import psycopg2
    from psycopg2.extras import execute_values, RealDictCursor
except ImportError:
    print("psycopg2 설치 필요: pip install psycopg2-binary")
    sys.exit(1)

from migration_config import DB_CONFIG, SCHEMA

# ============================================
# Configuration
# ============================================

EXCEL_FILE = Path(r"E:\workspace\idino_career\user_mig\idino_table_columns.xlsx")
SHEET_PREFIX = "○"
BATCH_SIZE = 1000

# UUID 변환용 네임스페이스 (고정값으로 결정론적 UUID 생성)
UUID_NAMESPACE = uuid.UUID('6ba7b810-9dad-11d1-80b4-00c04fd430c8')  # NAMESPACE_OID

# ============================================
# UUID 컬럼 목록 (테이블별)
# ============================================

UUID_COLUMNS = {
    "tb_activity": ["activity_id"],
    "tb_course_competency_map": ["map_id"],
    "tb_course_offering": ["offering_id"],
    "tb_cumulative_summary": ["summary_id"],
    "tb_enrollment": ["enrollment_id", "course_offering_id"],
    "tb_grade": ["grade_id", "enrollment_id"],
    "tb_grade_summary": ["summary_id"],
    "tb_professor_course": ["professor_course_id"],
    "tb_student_competency": ["student_competency_id"],
    "tb_user": ["user_id"],
    "tb_achievement": ["achievement_id"],
}

# ============================================
# DATE 컬럼 목록 (YYYYMMDD → DATE 변환)
# ============================================

DATE_COLUMNS = {
    "tb_program": ["start_date", "end_date"],
    "tb_activity": ["start_date", "end_date"],
}

# ============================================
# NOT NULL 기본값 설정
# ============================================

DEFAULT_VALUES = {
    "tb_student": {
        "admission_year": lambda row: extract_year_from_id(row.get("student_id")) or 2020,
        "current_grade": lambda row: safe_int(row.get("current_grade"), 1),
        "current_semester": lambda row: safe_int(row.get("current_semester"), 1),
    },
}

# ============================================
# 컬럼 매핑 (Excel → DB)
# ============================================

COLUMN_MAPPINGS = {
    "tb_enrollment": {
        "offering_id": "course_offering_id",
        "status": "status_cd",
    },
}

# ============================================
# FK 의존성 순서
# ============================================

MIGRATION_ORDER = [
    # Level 1-2: 기본 마스터
    ("tb_college", "○tb_college"),
    ("tb_competency", "○tb_competency"),
    ("tb_program", "○tb_program"),

    # Level 3
    ("tb_department", "○tb_department"),
    ("tb_course", "○tb_course"),

    # Level 4
    ("tb_professor", "○tb_professor"),
    ("tb_student", "○tb_student2"),

    # Level 5
    ("tb_course_offering", "○tb_course_offering"),
    ("tb_professor_course", "○tb_professor_course"),
    ("tb_user", "○tb_user"),

    # Level 6
    ("tb_enrollment", "○tb_enrollment"),
    ("tb_course_competency_map", "○tb_course_competency_map"),
    ("tb_student_competency", "○tb_student_competency"),
    ("tb_activity", "○tb_activity"),
    ("tb_achievement", "tb_achievement"),

    # Level 7
    ("tb_grade", "○tb_grade"),
    ("tb_grade_summary", "○tb_grade_summary"),
    ("tb_cumulative_summary", "○tb_cumulative_summary"),
]

EXCLUDE_COLUMN_PATTERNS = ["col_"]


# ============================================
# Helper Functions
# ============================================

def get_db_connection():
    """DB 연결"""
    return psycopg2.connect(**DB_CONFIG)


def to_uuid(value: Any, table_name: str = None) -> Optional[str]:
    """원본 값을 UUID로 변환 (결정론적)"""
    if value is None:
        return None
    # 이미 UUID 형식이면 그대로 반환
    str_val = str(value).strip()
    if not str_val:
        return None
    # UUID 패턴 체크
    uuid_pattern = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
    if uuid_pattern.match(str_val):
        return str_val
    # 결정론적 UUID 생성 (uuid5)
    return str(uuid.uuid5(UUID_NAMESPACE, str_val))


def to_date(value: Any) -> Optional[date]:
    """YYYYMMDD 정수를 DATE로 변환"""
    if value is None:
        return None
    try:
        if isinstance(value, (date, datetime)):
            return value if isinstance(value, date) else value.date()
        # 정수형 (20200522)
        if isinstance(value, (int, float)):
            str_val = str(int(value))
            if len(str_val) == 8:
                return date(int(str_val[:4]), int(str_val[4:6]), int(str_val[6:8]))
        # 문자열
        if isinstance(value, str):
            str_val = value.strip()
            if len(str_val) == 8 and str_val.isdigit():
                return date(int(str_val[:4]), int(str_val[4:6]), int(str_val[6:8]))
    except (ValueError, TypeError):
        pass
    return None


def extract_year_from_id(student_id: Any) -> Optional[int]:
    """학번에서 입학년도 추출 (예: 20201234 → 2020)"""
    if student_id is None:
        return None
    try:
        str_id = str(student_id).strip()
        if len(str_id) >= 4:
            year = int(str_id[:4])
            if 1990 <= year <= 2030:
                return year
    except (ValueError, TypeError):
        pass
    return 2020  # 기본값


def safe_int(value: Any, default: int = 0) -> int:
    """안전한 정수 변환"""
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value or value.upper() == "NULL":
                return default
        return int(float(value))
    except (ValueError, TypeError):
        return default


def should_exclude_column(col_name: str) -> bool:
    """컬럼 제외 여부"""
    if not col_name:
        return True
    for pattern in EXCLUDE_COLUMN_PATTERNS:
        if col_name.startswith(pattern):
            return True
    if "," in col_name:
        return True
    return False


def convert_value(value: Any) -> Any:
    """기본 값 변환"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value.lower() == "null":
            return None
        return value
    return value


def get_db_columns(table_name: str) -> List[str]:
    """DB 컬럼 목록"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
    """, (SCHEMA, table_name))
    columns = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()
    return columns


def get_pk_column(table_name: str) -> Optional[str]:
    """테이블의 PK 컬럼 조회"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.constraint_column_usage c ON tc.constraint_name = c.constraint_name
        WHERE tc.table_schema = %s AND tc.table_name = %s AND tc.constraint_type = 'PRIMARY KEY'
    """, (SCHEMA, table_name))
    result = cur.fetchone()
    cur.close()
    conn.close()
    return result[0] if result else None


# ============================================
# Data Transformation
# ============================================

def transform_row(row: Dict[str, Any], table_name: str) -> Dict[str, Any]:
    """행 데이터 변환"""
    transformed = {}

    uuid_cols = UUID_COLUMNS.get(table_name, [])
    date_cols = DATE_COLUMNS.get(table_name, [])
    defaults = DEFAULT_VALUES.get(table_name, {})

    for col, val in row.items():
        col_lower = col.lower()

        # UUID 변환
        if col_lower in uuid_cols:
            transformed[col_lower] = to_uuid(val, table_name)
        # DATE 변환
        elif col_lower in date_cols:
            transformed[col_lower] = to_date(val)
        else:
            transformed[col_lower] = convert_value(val)

    # 기본값 적용
    for col, default_fn in defaults.items():
        if col not in transformed or transformed[col] is None:
            transformed[col] = default_fn(transformed)

    return transformed


# ============================================
# Excel Reading
# ============================================

def read_excel_sheet(workbook, sheet_name: str, db_table: str) -> Generator[Dict[str, Any], None, None]:
    """Excel 시트 읽기 (변환 적용)"""
    if sheet_name not in workbook.sheetnames:
        print(f"  [WARN] 시트 없음: {sheet_name}")
        return

    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return

    header = rows[0]
    columns = [str(c).strip() if c else f"_empty_{i}" for i, c in enumerate(header)]
    col_mapping = COLUMN_MAPPINGS.get(db_table, {})

    for row in rows[1:]:
        row_dict = {}
        for i, col_name in enumerate(columns):
            if should_exclude_column(col_name):
                continue
            value = row[i] if i < len(row) else None
            mapped_name = col_mapping.get(col_name, col_name)
            row_dict[mapped_name] = value

        if all(v is None for v in row_dict.values()):
            continue

        # 데이터 변환 적용
        transformed = transform_row(row_dict, db_table)
        yield transformed


# ============================================
# Database Operations
# ============================================

def truncate_table(conn, table_name: str):
    """테이블 TRUNCATE"""
    cur = conn.cursor()
    try:
        cur.execute(f"TRUNCATE TABLE {SCHEMA}.{table_name} CASCADE")
        conn.commit()
        print(f"  [TRUNCATE] {table_name}")
    except Exception as e:
        conn.rollback()
        print(f"  [ERROR] TRUNCATE {table_name}: {e}")
        raise
    finally:
        cur.close()


def insert_batch_upsert(conn, table_name: str, columns: List[str], rows: List[Dict], pk_col: str = None) -> Tuple[int, int]:
    """배치 INSERT (ON CONFLICT DO NOTHING)"""
    if not rows:
        return 0, 0

    cur = conn.cursor()
    col_str = ", ".join(columns)
    placeholders = ", ".join(["%s"] * len(columns))

    # ON CONFLICT DO NOTHING (중복 무시)
    if pk_col and pk_col in columns:
        query = f"""
            INSERT INTO {SCHEMA}.{table_name} ({col_str})
            VALUES ({placeholders})
            ON CONFLICT ({pk_col}) DO NOTHING
        """
    else:
        query = f"INSERT INTO {SCHEMA}.{table_name} ({col_str}) VALUES ({placeholders})"

    values = []
    for row in rows:
        row_values = tuple(row.get(col) for col in columns)
        values.append(row_values)

    inserted = 0
    error_count = 0
    error_samples = []
    try:
        for val in values:
            try:
                cur.execute(query, val)
                inserted += cur.rowcount
            except Exception as e:
                error_count += 1
                if len(error_samples) < 3:
                    error_samples.append(str(e)[:200])
                conn.rollback()
                continue
        conn.commit()

        # 에러 로깅
        if error_samples:
            print(f"  [ERROR SAMPLES] {table_name}:")
            for i, err in enumerate(error_samples, 1):
                print(f"    {i}. {err}")

        return inserted, error_count
    except Exception as e:
        conn.rollback()
        print(f"  [ERROR] INSERT: {e}")
        raise
    finally:
        cur.close()


def get_insert_columns(excel_columns: List[str], db_columns: List[str]) -> List[str]:
    """공통 컬럼 결정"""
    excel_lower = {c.lower() for c in excel_columns if not should_exclude_column(c)}
    db_lower = {c.lower(): c for c in db_columns}
    return [db_lower[c] for c in excel_lower if c in db_lower]


# ============================================
# Migration
# ============================================

def migrate_table(workbook, db_table: str, excel_sheet: str, conn) -> Tuple[int, int]:
    """테이블 마이그레이션"""
    print(f"\n[MIGRATE] {db_table} <- {excel_sheet}")

    db_columns = get_db_columns(db_table)
    if not db_columns:
        print(f"  [ERROR] DB 테이블 없음")
        return 0, 0

    if excel_sheet not in workbook.sheetnames:
        print(f"  [SKIP] Excel 시트 없음")
        return 0, 0

    ws = workbook[excel_sheet]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        print(f"  [SKIP] 빈 시트")
        return 0, 0

    excel_columns = [str(c).strip() if c else f"_empty_{i}" for i, c in enumerate(first_row)]

    # 컬럼 매핑 적용
    col_mapping = COLUMN_MAPPINGS.get(db_table, {})
    mapped_columns = [col_mapping.get(c, c) for c in excel_columns]

    # 공통 컬럼
    insert_columns = get_insert_columns(mapped_columns, db_columns)
    if not insert_columns:
        print(f"  [ERROR] 공통 컬럼 없음")
        return 0, 0

    # PK 컬럼
    pk_col = get_pk_column(db_table)

    print(f"  DB: {len(db_columns)}개, Excel: {len([c for c in excel_columns if not should_exclude_column(c)])}개, 매핑: {len(insert_columns)}개")

    # TRUNCATE
    truncate_table(conn, db_table)

    # INSERT
    batch = []
    total_inserted = 0
    total_errors = 0
    total_rows = 0

    for row_dict in read_excel_sheet(workbook, excel_sheet, db_table):
        total_rows += 1
        batch.append(row_dict)

        if len(batch) >= BATCH_SIZE:
            inserted, errors = insert_batch_upsert(conn, db_table, insert_columns, batch, pk_col)
            total_inserted += inserted
            total_errors += errors
            print(f"  진행: {total_inserted:,}/{total_rows:,} (에러: {total_errors:,})    ", end="\r")
            batch = []

    if batch:
        inserted, errors = insert_batch_upsert(conn, db_table, insert_columns, batch, pk_col)
        total_inserted += inserted
        total_errors += errors

    print(f"  [DONE] {total_inserted:,}건 삽입 (총 {total_rows:,}행, 에러 {total_errors:,}건)    ")
    return total_inserted, total_rows


def verify_fk_integrity(conn) -> List[str]:
    """FK 무결성 검증"""
    print("\n" + "=" * 60)
    print("[VERIFY] FK 무결성 검증")
    print("=" * 60)

    issues = []
    fk_checks = [
        ("tb_department", "college_cd", "tb_college", "college_cd"),
        ("tb_course", "department_cd", "tb_department", "department_cd"),
        ("tb_student", "department_cd", "tb_department", "department_cd"),
        ("tb_professor", "department_cd", "tb_department", "department_cd"),
        ("tb_course_offering", "course_cd", "tb_course", "course_cd"),
        ("tb_enrollment", "student_id", "tb_student", "student_id"),
    ]

    cur = conn.cursor()
    for child_table, child_col, parent_table, parent_col in fk_checks:
        try:
            query = f"""
                SELECT COUNT(*) FROM {SCHEMA}.{child_table} c
                LEFT JOIN {SCHEMA}.{parent_table} p ON c.{child_col}::text = p.{parent_col}::text
                WHERE c.{child_col} IS NOT NULL AND p.{parent_col} IS NULL
            """
            cur.execute(query)
            orphan_count = cur.fetchone()[0]

            if orphan_count > 0:
                issues.append(f"{child_table}.{child_col}: {orphan_count}건 참조 누락")
                print(f"  [WARN] {child_table}.{child_col} -> {parent_table}: {orphan_count}건 누락")
            else:
                print(f"  [OK] {child_table}.{child_col} -> {parent_table}")
        except Exception as e:
            print(f"  [SKIP] {child_table}: {e}")

    cur.close()
    return issues


def print_row_counts(conn):
    """결과 출력"""
    print("\n" + "=" * 60)
    print("[RESULT] 마이그레이션 결과")
    print("=" * 60)

    cur = conn.cursor()
    print(f"{'테이블':<30} {'행 수':>12}")
    print("-" * 45)

    total = 0
    for db_table, _ in MIGRATION_ORDER:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {SCHEMA}.{db_table}")
            count = cur.fetchone()[0]
            total += count
            print(f"{db_table:<30} {count:>12,}")
        except Exception as e:
            print(f"{db_table:<30} {'ERROR':>12}")

    print("-" * 45)
    print(f"{'합계':<30} {total:>12,}")
    cur.close()


def main():
    """메인"""
    print("=" * 60)
    print("IDINO Career - Excel 마이그레이션 V2 (데이터 변환 포함)")
    print("=" * 60)

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel 파일 없음: {EXCEL_FILE}")
        sys.exit(1)

    print(f"Excel: {EXCEL_FILE}")
    print(f"스키마: {SCHEMA}")
    print(f"테이블: {len(MIGRATION_ORDER)}개")

    print("\nExcel 로드 중...")
    workbook = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    print(f"시트: {len(workbook.sheetnames)}개")

    conn = get_db_connection()
    conn.autocommit = False

    try:
        # FK 비활성화
        cur = conn.cursor()
        cur.execute("SET session_replication_role = 'replica';")
        conn.commit()
        cur.close()
        print("\n[INFO] FK 제약조건 비활성화")

        # 마이그레이션
        for db_table, excel_sheet in MIGRATION_ORDER:
            try:
                migrate_table(workbook, db_table, excel_sheet, conn)
            except Exception as e:
                print(f"  [ERROR] {db_table}: {e}")

        # FK 활성화
        cur = conn.cursor()
        cur.execute("SET session_replication_role = 'origin';")
        conn.commit()
        cur.close()
        print("\n[INFO] FK 제약조건 활성화")

        print_row_counts(conn)
        verify_fk_integrity(conn)

    except Exception as e:
        conn.rollback()
        print(f"\n[FATAL] {e}")
        raise
    finally:
        workbook.close()
        conn.close()

    print("\n" + "=" * 60)
    print("마이그레이션 완료")
    print("=" * 60)


if __name__ == "__main__":
    main()
